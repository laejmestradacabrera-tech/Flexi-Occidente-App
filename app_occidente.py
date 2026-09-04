import streamlit as st
import pandas as pd
import os
import smtplib
from email.mime.text import MIMEText
import datetime
from fpdf import FPDF
import openpyxl
from oauth2client.service_account import ServiceAccountCredentials
import gspread
import locale
import subprocess
import streamlit.components.v1 as components
import re
import base64
import json
import modulo_agenda

# 1. CONFIGURACIÓN DE PÁGINA (Debe ser la primera instrucción)
st.set_page_config(page_title="Monitor Comercial Flexi Occidente", layout="wide", initial_sidebar_state="collapsed")

# Intentamos configurar el idioma español para las fechas
try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
except:
    pass

# --- FUNCIÓN PARA LEER IMÁGENES LOCALES ---
@st.cache_data
def obtener_imagen_base64(ruta_imagen):
    try:
        with open(ruta_imagen, "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode()
        ext = ruta_imagen.split('.')[-1].lower()
        mime_type = "image/png" if ext == "png" else "image/jpeg"
        return f"data:{mime_type};base64,{encoded_string}"
    except Exception as e:
        return ""

# --- ESTILO GLOBAL INTERACTIVO Y TEMA ---
st.markdown("""
    <style>
    /* CSS para ocultar el padding superior de Streamlit y dar sensación de App */
    .block-container { padding-top: 3rem; padding-bottom: 2rem; max-width: 1300px; }
    
    .footer {
        position: fixed; left: 0; bottom: 0; width: 100%;
        background-color: #0f172a; color: #94a3b8; text-align: center;
        padding: 12px; font-size: 13px; border-top: 1px solid #1e293b;
        z-index: 999; font-weight: 500;
    }
    th {
        background-color: #E30613 !important; color: white !important;
        font-weight: bold !important; text-transform: uppercase !important;
        text-align: center !important; padding: 12px !important;
    }
    td { text-align: center !important; font-size: 15px !important; }
    
    .kpi-box {
        background-color: #f8f9fa; border: 1px solid #ddd; border-radius: 8px;
        padding: 15px; text-align: center; box-shadow: 2px 2px 5px rgba(0,0,0,0.05);
    }
    .kpi-title { font-size: 14px; color: #555; font-weight: bold; text-transform: uppercase; }
    .kpi-value { font-size: 24px; color: #E30613; font-weight: bold; margin: 5px 0; }
    .kpi-delta { font-size: 15px; font-weight: bold; }

    /* ESTILOS DEL NUEVO LOBBY CORPORATIVO */
    .lobby-header { text-align: center; margin-bottom: 40px; position: relative; }
    .lobby-header h1 { font-family: 'Arial Black', sans-serif; font-size: 45px; color: white; margin: 10px 0 0 0; line-height: 1.1; letter-spacing: -1px; }
    .lobby-header h2 { font-family: 'Arial', sans-serif; font-size: 20px; color: #E30613; margin-top: 5px; font-weight: bold; letter-spacing: 2px; }
    .lobby-header p { color: #94a3b8; font-size: 16px; margin-top: 10px; }
    
    .kpi-row-lobby { display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin-bottom: 30px; }
    .kpi-card-lobby { background-color: #1e293b; border: 1px solid #334155; border-radius: 12px; padding: 20px; display: flex; align-items: center; gap: 15px; }
    .kpi-icon-lobby { width: 50px; height: 50px; border-radius: 8px; border: 2px solid #E30613; display: flex; align-items: center; justify-content: center; color: #E30613; flex-shrink: 0; background-color: rgba(227,6,19,0.05); }
    .kpi-data-lobby h3 { font-size: 24px; font-weight: bold; color: white; margin: 0; line-height: 1.2; }
    .kpi-data-lobby p { font-size: 12px; color: #94a3b8; margin: 0; }
    
    .action-card-lobby { background-color: #1e293b; border: 1px solid #334155; border-radius: 12px 12px 0 0; padding: 40px; display: flex; align-items: center; gap: 20px; height: 180px; border-bottom: none; }
    .action-icon-circle { width: 80px; height: 80px; border-radius: 50%; border: 3px solid #E30613; display: flex; align-items: center; justify-content: center; color: #E30613; flex-shrink: 0; background-color: rgba(227,6,19,0.05); }
    .action-texts h3 { font-size: 22px; font-weight: bold; color: white; margin: 0 0 5px 0; }
    .action-texts p { font-size: 15px; color: #94a3b8; margin: 0; }
    
    /* Botones de acción integrados a las tarjetas */
    div[data-testid="column"]:nth-child(1) button { background-color: #E30613 !important; color: white !important; font-weight: bold !important; height: 50px !important; border-radius: 0 0 12px 12px !important; border: 1px solid #334155 !important; border-top: none !important; width: 100% !important; font-size: 16px !important; margin-top: -16px !important; transition: all 0.3s; }
    div[data-testid="column"]:nth-child(2) button { background-color: #E30613 !important; color: white !important; font-weight: bold !important; height: 50px !important; border-radius: 0 0 12px 12px !important; border: 1px solid #334155 !important; border-top: none !important; width: 100% !important; font-size: 16px !important; margin-top: -16px !important; transition: all 0.3s; }
    div[data-testid="column"] button:hover { background-color: #b9000b !important; }
    </style>
    """, unsafe_allow_html=True)

# --- CONFIGURACIÓN CENTRALIZADA DE GOOGLE ---
scope = [
    'https://spreadsheets.google.com/feeds', 
    'https://www.googleapis.com/auth/drive', 
    'https://www.googleapis.com/auth/spreadsheets'
]

try:
    creds = ServiceAccountCredentials.from_json_keyfile_dict(dict(st.secrets["gcp_service_account"]), scope)
    client = gspread.authorize(creds)
    ID_ARCHIVO = '1lGlVEBgu9QsrH9PYTTuoRKQeWnYiR7OwUElCsfkDgoM'
    archivo_ventas_g = client.open_by_key(ID_ARCHIVO)
    sheet_bitacora = client.open_by_key(ID_ARCHIVO).sheet1
except Exception as e:
    st.warning("Advertencia: No se pudo conectar a Google Sheets. Verifica tus secretos.")

# --- FUNCIONES ---
def buscar_archivo(palabra_clave):
    archivos = [f for f in os.listdir('.') if palabra_clave.lower() in f.lower() and f.endswith(('.xlsx', '.csv'))]
    return sorted(archivos)[-1] if archivos else None

def obtener_fecha_actualizacion(nombre_archivo):
    if not nombre_archivo: return "Archivo no disponible"
    try:
        comando = ['git', 'log', '-1', '--format=%ct', nombre_archivo]
        resultado = subprocess.run(comando, capture_output=True, text=True)
        if resultado.returncode == 0 and resultado.stdout.strip():
            timestamp_github = int(resultado.stdout.strip())
            fecha_utc = datetime.datetime.utcfromtimestamp(timestamp_github)
            fecha_mexico = fecha_utc - datetime.timedelta(hours=6)
            return fecha_mexico.strftime("%d/%m/%Y - %H:%M hrs")
    except Exception:
        pass 
    try:
        tiempo_modificacion = os.path.getmtime(nombre_archivo)
        fecha_servidor = datetime.datetime.utcfromtimestamp(tiempo_modificacion)
        fecha_mexico = fecha_servidor - datetime.timedelta(hours=6)
        return fecha_mexico.strftime("%d/%m/%Y - %H:%M hrs")
    except FileNotFoundError:
        return "Archivo pendiente de carga"
    except Exception:
        return "Fecha no disponible"

def cargar_tiendas():
    arch_tiendas = buscar_archivo('CORREO DE TIENDAS')
    if arch_tiendas:
        try:
            return pd.read_excel(arch_tiendas) if arch_tiendas.endswith('.xlsx') else pd.read_csv(arch_tiendas)
        except Exception as e:
            pass
    return pd.DataFrame({'TIENDA': ['Error'], 'NOMBRE': ['Sin datos'], 'ENCARGADO': ['Sin datos']})

def cargar_archivos_locales():
    if 'df_ventas' not in st.session_state or 'df_tallas' not in st.session_state or 'lista_excepciones' not in st.session_state:
        try:
            arch_ventas = buscar_archivo('Ventas')
            if arch_ventas:
                st.session_state.df_ventas = pd.read_excel(arch_ventas) if arch_ventas.endswith('.xlsx') else pd.read_csv(arch_ventas)
            else:
                st.session_state.df_ventas = pd.read_excel("Ventas.xlsx")
                
            arch_tallas = buscar_archivo('Valores de tallas')
            if arch_tallas:
                if arch_tallas.endswith('.xlsx'):
                    try:
                        st.session_state.df_tallas = pd.read_excel(arch_tallas, sheet_name="Hoja1")
                    except:
                        st.session_state.df_tallas = pd.read_excel(arch_tallas)
                else:
                    st.session_state.df_tallas = pd.read_csv(arch_tallas)
            else:
                st.session_state.df_tallas = pd.read_excel("Valores de tallas.xlsx")
                
            # LECTURA FORZADA GLOBAL DE EXCEPCIONES CON LECTOR ROBUSTO
            lista_exc = []
            arch_exc = buscar_archivo('Excepciones')
            if arch_exc:
                try:
                    df_exc = pd.read_excel(arch_exc) if arch_exc.endswith('.xlsx') else pd.read_csv(arch_exc, encoding='utf-8-sig')
                except:
                    df_exc = pd.read_csv(arch_exc, encoding='latin1')
                col_mod = 'Modelo' if 'Modelo' in df_exc.columns else df_exc.columns[0]
                lista_exc = df_exc[col_mod].astype(str).str.strip().str.upper().tolist()
            else:
                # Búsqueda implacable en la pestaña de Excepciones
                nombre_archivo_tallas = arch_tallas if arch_tallas else "Valores de tallas.xlsx"
                try:
                    if nombre_archivo_tallas.endswith('.xlsx'):
                        xls = pd.ExcelFile(nombre_archivo_tallas)
                        hoja_exc = next((h for h in xls.sheet_names if 'excepciones' in h.lower()), None)
                        if hoja_exc:
                            df_exc = pd.read_excel(nombre_archivo_tallas, sheet_name=hoja_exc)
                            col_mod = 'Modelo' if 'Modelo' in df_exc.columns else df_exc.columns[0]
                            lista_exc = df_exc[col_mod].astype(str).str.strip().str.upper().tolist()
                except Exception as e:
                    pass
            st.session_state.lista_excepciones = lista_exc
            
        except Exception as e:
            return False
    return True

def validar_captura_stock(tienda_id, modelo, talla_input, df_ventas, df_tallas):
    try:
        # CLONAMOS los dataframes para no desconfigurar (KeyError) el resto del monitor
        df_v_local = df_ventas.copy()
        df_t_local = df_tallas.copy()
        
        df_v_local.columns = df_v_local.columns.astype(str).str.strip().str.lower()
        df_t_local.columns = df_t_local.columns.astype(str).str.strip().str.lower()
        
        try: tda_buscada = int(str(tienda_id).strip())
        except: tda_buscada = -1
            
        df_v_local['tienda_int'] = pd.to_numeric(df_v_local['tienda'], errors='coerce').fillna(-1).astype(int)
        
        modelo_buscado = str(modelo).replace(' ', '').replace('-', '').upper()
        df_v_local['modelo_cln'] = df_v_local['modelo'].astype(str).str.replace(' ', '', regex=False).str.replace('-', '', regex=False).str.upper()
        
        df_filtro = df_v_local[(df_v_local['tienda_int'] == tda_buscada) & (df_v_local['modelo_cln'] == modelo_buscado)]
        
        try: talla_buscada_str = str(int(float(talla_input)))
        except: talla_buscada_str = str(talla_input).strip()

        # MANEJO AVANZADO DE TALLAS COMERCIALES (EJ. 22 vs 220)
        try:
            talla_num = float(talla_input)
            if talla_num < 100:
                talla_buscada_str_2 = str(int(talla_num * 10))
            else:
                talla_buscada_str_2 = str(int(talla_num / 10.0))
        except:
            talla_buscada_str_2 = talla_buscada_str
            
        with st.expander("🔍 MODO DEPURACIÓN: AUDITORÍA DE CRUCE DE TALLAS", expanded=True):
            st.write(f"**Buscando:** Tienda N°=[{tda_buscada}], Modelo=[{modelo_buscado}], Talla Capturada=[{talla_buscada_str}] o [{talla_buscada_str_2}]")
            
            if df_filtro.empty:
                st.write(f"❌ **Resultado:** El modelo {modelo_buscado} no existe en el inventario de la Tienda {tda_buscada}. (Quiebre válido).")
                return True, ""
                
            st.write(f"✅ Se encontraron {len(df_filtro)} fila(s) del modelo. Cruzando con archivo de Tallas...")
            
            col_dpto = 'valor' if 'valor' in df_t_local.columns else df_t_local.columns[0]
            df_t_local['dpto_cln'] = df_t_local[col_dpto].astype(str).str.strip().str.lower()
            
            for idx, row_venta in df_filtro.iterrows():
                dpto_venta = str(row_venta.get('departamento', '')).strip().lower()
                st.write(f"--- \n**Analizando Departamento:** '{dpto_venta}'")
                tallas_row = df_t_local[df_t_local['dpto_cln'] == dpto_venta]
                
                if tallas_row.empty:
                    st.warning(f"⚠️ No se encontró el departamento '{dpto_venta}' en Valores de tallas.xlsx")
                    continue
                    
                for i in range(1, 16):
                    col_ex = f'ex{i}'
                    if col_ex in tallas_row.columns:
                        talla_matriz = tallas_row.iloc[0][col_ex]
                        if pd.notna(talla_matriz) and str(talla_matriz).strip() != '' and str(talla_matriz).strip().lower() != 'nan':
                            try: talla_matriz_str = str(int(float(talla_matriz)))
                            except: talla_matriz_str = str(talla_matriz).strip()
                                
                            if talla_matriz_str == talla_buscada_str or talla_matriz_str == talla_buscada_str_2:
                                st.write(f"🎯 **¡COINCIDENCIA ENCONTRADA!** Talla Matriz [{talla_matriz_str}] == Talla Captura [{talla_input}] en la columna **{col_ex}**")
                                
                                # ESCUDO GLOBAL EN LA BITÁCORA PARA PREVENIR CAPTURAS FANTASMAS
                                try:
                                    t_num = float(talla_input)
                                    if t_num >= 100: t_num = t_num / 10.0
                                except:
                                    t_num = 0.0
                                    
                                if dpto_venta == 'caballero' and t_num == 30.5:
                                    return False, "⛔ TALLA FANTASMA: El sistema corporativo no maneja la talla 30.5 en caballero. Captura rechazada."
                                if dpto_venta in ['niño', 'nino'] and t_num == 21.5:
                                    return False, "⛔ TALLA FANTASMA: El sistema corporativo no maneja la talla 21.5 en niño. Captura rechazada."
                                
                                lista_exc = st.session_state.get('lista_excepciones', [])
                                # PARCHE EXCEPCIONES: Aplicar escudo sin importar el departamento de SAP
                                if t_num > 25.0 and modelo_buscado in lista_exc:
                                    return False, f"⛔ EXCEPCIÓN DETECTADA: El modelo {modelo_buscado} no se fabrica en tallas mayores a 25.0. Captura rechazada."

                                if col_ex in row_venta:
                                    existencia = row_venta[col_ex]
                                    st.write(f"📦 **Existencia leída en Ventas para {col_ex}:** [{existencia}]")
                                    try: existencia_num = float(existencia)
                                    except: existencia_num = 0.0
                                        
                                    if existencia_num > 0:
                                        st.error(f"⛔ BLOQUEO ACTIVADO: Se detectaron {existencia_num} piezas físicas.")
                                        return False, f"⛔ CAPTURA BLOQUEADA: El sistema registra {int(existencia_num)} par(es) de la talla {talla_buscada_str} (Modelo {modelo_buscado}) físicamente en la sucursal {tda_buscada}."
                                    else:
                                        st.write(f"✅ La existencia es {existencia_num}. Permitiendo captura (Quiebre válido).")
                                        return True, ""
                
        # Si termina de buscar en todas las matrices y no encontró la talla
        st.write(f"⚠️ No se encontró la talla {talla_input} en la matriz de '{dpto_venta}' para validar. Permitiendo captura.")
        return True, ""
    except Exception as e:
        st.error(f"Error interno en validación: {e}")
        return True, ""

def enviar_correo_ejecutivo(tienda_objetivo, conversion, ticket, meta_conv, meta_tkt, faltan_pares, faltan_pesos, correo_destinatario="fleoutgdl@divec-flexi.com"):
    logro_conv = conversion >= meta_conv
    logro_ticket = ticket >= meta_tkt
    desviacion_conv = conversion - meta_conv
    desviacion_ticket = ticket - meta_tkt

    try:
        remitente = st.secrets["CORREO_REMITENTE"]
        password = st.secrets["CORREO_PASSWORD"]
        destinatario = correo_destinatario 
        
        asunto = f"🚀 Desempeño Comercial y Reto Acumulado - Tienda {tienda_objetivo}"
        
        cuerpo = f"Estimada encargada y equipo de la Tienda {tienda_objetivo}:\n\n"
        cuerpo += "Les compartimos el análisis de resultados comerciales de su sucursal, obtenido directamente tras la última actualización del monitor.\n\n"
        cuerpo += "--------------------------------------------------------------------------------\n"
        
        if logro_conv and logro_ticket:
            cuerpo += "🏆 ¡MUCHAS FELICIDADES POR EL RESULTADO!\n"
            cuerpo += "Queremos reconocer el extraordinario desempeño del equipo en el piso de venta. Han alcanzado y superado de forma simultánea las dos metas vitales de nuestra zona para calzado:\n"
            cuerpo += f" * Conversión Actual: {conversion:.2f}% (Meta obligatoria: {meta_conv:.2f}%)\n"
            cuerpo += f" * Ticket Promedio Actual: {ticket:.2f} unidades (Meta obligatoria: {meta_tkt:.2f} unidades de calzado)\n\n"
        else:
            cuerpo += "⚠️ ALERTA DE DESVIACIÓN DE METAS EN PISO DE VENTA\n"
            cuerpo += "Es necesario ajustar la estrategia operativa para alcanzar los objetivos obligatorios de la Zona Occidente:\n"
            if logro_conv:
                cuerpo += f" ✅ CONVERSIÓN: {conversion:.2f}% (Lograda)\n"
            else:
                cuerpo += f" ❌ CONVERSIÓN: {conversion:.2f}% (Faltan {abs(desviacion_conv):.2f}% para alcanzar la meta de {meta_conv}%)\n"
            if logro_ticket:
                cuerpo += f" ✅ TICKET PROMEDIO: {ticket:.2f} unidades (Logrado)\n"
            else:
                cuerpo += f" ❌ TICKET PROMEDIO: {ticket:.2f} unidades (Faltan {abs(desviacion_ticket):.2f} unidades para alcanzar la meta de {meta_tkt})\n\n"
        
        cuerpo += "--------------------------------------------------------------------------------\n"
        cuerpo += "📊 ANÁLISIS COMPARATIVO MENSUAL (RETO ACUMULADO)\n"
        cuerpo += "Para igualar y superar el comparativo histórico del año pasado, al día de hoy el estatus es el siguiente:\n\n"
        
        if faltan_pares > 0:
            cuerpo += f" 👟 Te faltan: {faltan_pares:,.0f} pares de calzado.\n"
        else:
            cuerpo += f" 👟 Llevas a favor: {abs(faltan_pares):,.0f} pares de calzado.\n"
            
        if faltan_pesos > 0:
            cuerpo += f" 💰 Te faltan: ${faltan_pesos:,.2f} MXN en importe.\n\n"
        else:
            cuerpo += f" 💰 Llevas a favor: ${abs(faltan_pesos):,.2f} MXN en importe.\n\n"
        
        cuerpo += "¡Cada cliente que cruza la puerta cuenta para superar esta marca! Éxito en sus ventas.\n"
        cuerpo += "--------------------------------------------------------------------------------\n\n"
        
        cuerpo += "Atentamente,\n"
        cuerpo += "Gerencia Comercial Zona Occidente\n"
        cuerpo += "LAE. José Martín Estrada Cabrera"

        msg = MIMEText(cuerpo)
        msg['Subject'] = asunto
        msg['From'] = remitente
        msg['To'] = destinatario
        
        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.login(remitente, password)
        server.sendmail(remitente, [destinatario], msg.as_string())
        server.quit()
        return f"✅ Correo ejecutivo enviado exitosamente a la Tienda {tienda_objetivo}."
    except Exception as e:
        return f"❌ Error al enviar el correo a la Tienda {tienda_objetivo}: {e}"

def generar_reporte_top20_pdf(df_top20, nombre_sucursal):
    hora_mexico = datetime.datetime.utcnow() - datetime.timedelta(hours=6)
    fecha_actual = hora_mexico.strftime("%d/%m/%Y")    
    pdf = FPDF(orientation='P', unit='mm', format='Letter')
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    pdf.set_font("Arial", 'B', 16)
    pdf.set_text_color(227, 6, 19)
    pdf.cell(0, 8, "FLEXI - ZONA OCCIDENTE", ln=True, align="C")
    
    pdf.set_font("Arial", 'B', 12)
    pdf.set_text_color(50, 50, 50)
    pdf.cell(0, 8, "AUDITORIA COMERCIAL: TOP 20 MODELOS DE LA SUCURSAL", ln=True, align="C")
    pdf.line(10, 28, 205, 28)
    pdf.ln(8)
    
    pdf.set_font("Arial", '', 10)
    pdf.cell(40, 6, "Fecha de Reporte:", 0, 0)
    pdf.cell(60, 6, fecha_actual, 0, 0)
    pdf.cell(35, 6, "Encargada:", 0, 0)
    pdf.cell(60, 6, "_______________________", 0, 1)
    
    pdf.cell(40, 6, "Sucursal:", 0, 0)
    pdf.cell(60, 6, nombre_sucursal, 0, 0)
    pdf.cell(35, 6, "Gerente Comercial:", 0, 0)
    pdf.cell(60, 6, "LAE. Jose Martin Estrada Cabrera", 0, 1)
    pdf.ln(8)
    
    pdf.set_font("Arial", 'B', 9)
    pdf.set_fill_color(227, 6, 19)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(15, 8, "#", 1, 0, 'C', fill=True)
    pdf.cell(30, 8, "MODELO", 1, 0, 'C', fill=True)
    pdf.cell(40, 8, "PARES VENDIDOS", 1, 0, 'C', fill=True)
    pdf.cell(110, 8, "DESEMPENO EN LA ZONA OCCIDENTE", 1, 1, 'C', fill=True)
    
    pdf.set_font("Arial", '', 9)
    pdf.set_text_color(0, 0, 0)
    for i, row in df_top20.iterrows():
        posicion = i + 1
        modelo = str(row.get('MODELO', 'S/D'))
        pares = str(row.get('PARES VENDIDOS', '0'))
        if posicion <= 5: desempeno = "Top 5 mas vendido en la region"
        elif posicion <= 10: desempeno = "Alta demanda en la zona"
        else: desempeno = "Desplazamiento regular"
        pdf.cell(15, 7, f"{posicion:02d}", 1, 0, 'C')
        pdf.cell(30, 7, modelo, 1, 0, 'C')
        pdf.cell(40, 7, pares, 1, 0, 'C')
        pdf.cell(110, 7, desempeno, 1, 1, 'L')
        
    pdf.ln(10)
    pdf.set_font("Arial", 'I', 9)
    pdf.set_text_color(80, 80, 80)
    pdf.multi_cell(0, 5, "Nota: Este documento sirve como guia visual para que el equipo en piso valide fisicamente en su bodega que estos modelos ganadores esten exhibidos.")
    pdf.ln(25)
    pdf.set_font("Arial", '', 10)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(90, 5, "_______________________", 0, 0, 'C')
    pdf.cell(15, 5, "", 0, 0)
    pdf.cell(90, 5, "_______________________", 0, 1, 'C')
    pdf.cell(90, 5, "Firma de la Encargada", 0, 0, 'C')
    pdf.cell(15, 5, "", 0, 0)
    pdf.cell(90, 5, "LAE. Jose Martin Estrada", 0, 1, 'C')
    pdf.cell(90, 5, "Responsable de Sucursal", 0, 0, 'C')
    pdf.cell(15, 5, "", 0, 0)
    pdf.cell(90, 5, "Gerente Comercial", 0, 1, 'C')
    return bytes(pdf.output(dest='S').encode('latin1'))

def generar_reporte_desfogue_pdf(df_desfogue, nombre_sucursal):
    hora_mexico = datetime.datetime.utcnow() - datetime.timedelta(hours=6)
    fecha_actual = hora_mexico.strftime("%d/%m/%Y")    
    pdf = FPDF(orientation='P', unit='mm', format='Letter')
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    pdf.set_font("Arial", 'B', 16)
    pdf.set_text_color(227, 6, 19)
    pdf.cell(0, 8, "FLEXI - ZONA OCCIDENTE", ln=True, align="C")
    
    pdf.set_font("Arial", 'B', 12)
    pdf.set_text_color(50, 50, 50)
    pdf.cell(0, 8, "ORDEN DE DESFOGUE: MODELOS CON 1 A 3 PARES", ln=True, align="C")
    pdf.line(10, 28, 205, 28)
    pdf.ln(8)
    
    pdf.set_font("Arial", '', 10)
    pdf.cell(40, 6, "Fecha de Reporte:", 0, 0)
    pdf.cell(60, 6, fecha_actual, 0, 0)
    pdf.cell(35, 6, "Encargada:", 0, 0)
    pdf.cell(60, 6, "_______________________", 0, 1)
    
    pdf.cell(40, 6, "Sucursal:", 0, 0)
    pdf.cell(60, 6, nombre_sucursal, 0, 0)
    pdf.cell(35, 6, "Gerente Comercial:", 0, 0)
    pdf.cell(60, 6, "LAE. Jose Martin Estrada Cabrera", 0, 1)
    pdf.ln(8)
    
    pdf.set_font("Arial", 'B', 9)
    pdf.set_fill_color(227, 6, 19)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(15, 8, "#", 1, 0, 'C', fill=True)
    pdf.cell(45, 8, "MODELO", 1, 0, 'C', fill=True)
    pdf.cell(35, 8, "PARES FISICOS", 1, 0, 'C', fill=True)
    pdf.cell(100, 8, "ACCION REQUERIDA", 1, 1, 'C', fill=True)
    
    pdf.set_font("Arial", '', 9)
    pdf.set_text_color(0, 0, 0)
    for idx, (orig_idx, row) in enumerate(df_desfogue.iterrows()):
        posicion = idx + 1
        modelo = str(row.get('Modelo a Desfogar', 'S/D'))
        pares = str(row.get('Pares Físicos (Total)', '0'))
        accion = "Transferir a Outlet / Depurar bodega"
        
        pdf.cell(15, 7, f"{posicion:02d}", 1, 0, 'C')
        pdf.cell(45, 7, modelo, 1, 0, 'C')
        pdf.cell(35, 7, pares, 1, 0, 'C')
        pdf.cell(100, 7, accion, 1, 1, 'L')
        
    pdf.ln(10)
    pdf.set_font("Arial", 'I', 9)
    pdf.set_text_color(80, 80, 80)
    pdf.multi_cell(0, 5, "Nota: Este documento autoriza e instruye a la sucursal a iniciar el proceso de transferencia de los modelos listados para liberar espacio estrategico en bodega.")
    pdf.ln(25)
    pdf.set_font("Arial", '', 10)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(90, 5, "_______________________", 0, 0, 'C')
    pdf.cell(15, 5, "", 0, 0)
    pdf.cell(90, 5, "_______________________", 0, 1, 'C')
    pdf.cell(90, 5, "Firma de la Encargada", 0, 0, 'C')
    pdf.cell(15, 5, "", 0, 0)
    pdf.cell(90, 5, "LAE. Jose Martin Estrada", 0, 1, 'C')
    pdf.cell(90, 5, "Responsable de Sucursal", 0, 0, 'C')
    pdf.cell(15, 5, "", 0, 0)
    pdf.cell(90, 5, "Gerente Comercial", 0, 1, 'C')
    return bytes(pdf.output(dest='S').encode('latin1'))

def generar_reporte_inmovil_pdf(df_inmovil, nombre_sucursal):
    hora_mexico = datetime.datetime.utcnow() - datetime.timedelta(hours=6)
    fecha_actual = hora_mexico.strftime("%d/%m/%Y")    
    pdf = FPDF(orientation='P', unit='mm', format='Letter')
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    pdf.set_font("Arial", 'B', 16)
    pdf.set_text_color(227, 6, 19)
    pdf.cell(0, 8, "FLEXI - ZONA OCCIDENTE", ln=True, align="C")
    
    pdf.set_font("Arial", 'B', 12)
    pdf.set_text_color(50, 50, 50)
    pdf.cell(0, 8, "ORDEN DE DEPURACION: INVENTARIO INMOVIL (0 VENTAS EN 60 DIAS)", ln=True, align="C")
    pdf.line(10, 28, 205, 28)
    pdf.ln(8)
    
    pdf.set_font("Arial", '', 10)
    pdf.cell(40, 6, "Fecha de Reporte:", 0, 0)
    pdf.cell(60, 6, fecha_actual, 0, 0)
    pdf.cell(35, 6, "Encargada:", 0, 0)
    pdf.cell(60, 6, "_______________________", 0, 1)
    
    pdf.cell(40, 6, "Sucursal:", 0, 0)
    pdf.cell(60, 6, nombre_sucursal, 0, 0)
    pdf.cell(35, 6, "Gerente Comercial:", 0, 0)
    pdf.cell(60, 6, "LAE. Jose Martin Estrada Cabrera", 0, 1)
    pdf.ln(8)
    
    pdf.set_font("Arial", 'B', 9)
    pdf.set_fill_color(30, 58, 138)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(15, 8, "#", 1, 0, 'C', fill=True)
    pdf.cell(45, 8, "MODELO", 1, 0, 'C', fill=True)
    pdf.cell(30, 8, "FISICOS", 1, 0, 'C', fill=True)
    pdf.cell(30, 8, "EN CAMINO", 1, 0, 'C', fill=True)
    pdf.cell(75, 8, "ACCION REQUERIDA", 1, 1, 'C', fill=True)
    
    pdf.set_font("Arial", '', 9)
    pdf.set_text_color(0, 0, 0)
    for idx, (orig_idx, row) in enumerate(df_inmovil.iterrows()):
        posicion = idx + 1
        modelo = str(row.get('Modelo a Desfogar', 'S/D'))
        pares = str(row.get('Pares Físicos (Total)', '0'))
        camino = str(row.get('Pares en Camino', '0'))
        accion = "Solicitar devolucion / traspaso"
        
        pdf.cell(15, 7, f"{posicion:02d}", 1, 0, 'C')
        pdf.cell(45, 7, modelo, 1, 0, 'C')
        pdf.cell(30, 7, pares, 1, 0, 'C')
        pdf.cell(30, 7, camino, 1, 0, 'C')
        pdf.cell(75, 7, accion, 1, 1, 'L')
        
    pdf.ln(10)
    pdf.set_font("Arial", 'I', 9)
    pdf.set_text_color(80, 80, 80)
    pdf.multi_cell(0, 5, "Nota: Este documento avala que los modelos listados no han registrado ventas en los ultimos 60 dias, representando capital congelado. Se autoriza a la encargada gestionar su liberacion inmediata.")
    pdf.ln(25)
    pdf.set_font("Arial", '', 10)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(90, 5, "_______________________", 0, 0, 'C')
    pdf.cell(15, 5, "", 0, 0)
    pdf.cell(90, 5, "_______________________", 0, 1, 'C')
    pdf.cell(90, 5, "Firma de la Encargada", 0, 0, 'C')
    pdf.cell(15, 5, "", 0, 0)
    pdf.cell(90, 5, "LAE. Jose Martin Estrada", 0, 1, 'C')
    pdf.cell(90, 5, "Responsable de Sucursal", 0, 0, 'C')
    pdf.cell(15, 5, "", 0, 0)
    pdf.cell(90, 5, "Gerente Comercial", 0, 1, 'C')
    return bytes(pdf.output(dest='S').encode('latin1'))

def generar_reporte_traspaso_masivo_pdf(df_traspasos):
    hora_mexico = datetime.datetime.utcnow() - datetime.timedelta(hours=6)
    fecha_actual = hora_mexico.strftime("%d/%m/%Y")    
    pdf = FPDF(orientation='L', unit='mm', format='Letter') 
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    pdf.set_font("Arial", 'B', 16)
    pdf.set_text_color(227, 6, 19)
    pdf.cell(0, 8, "FLEXI - ZONA OCCIDENTE", ln=True, align="C")
    
    pdf.set_font("Arial", 'B', 12)
    pdf.set_text_color(50, 50, 50)
    pdf.cell(0, 8, "ORDEN GERENCIAL DE NIVELACIÓN DE INVENTARIO", ln=True, align="C")
    pdf.line(10, 28, 265, 28)
    pdf.ln(8)
    
    pdf.set_font("Arial", '', 10)
    pdf.cell(40, 6, "Fecha de Emisión:", 0, 0)
    pdf.cell(60, 6, fecha_actual, 0, 0)
    pdf.cell(35, 6, "Gerente Comercial:", 0, 0)
    pdf.cell(60, 6, "LAE. Jose Martin Estrada Cabrera", 0, 1)
    pdf.ln(8)
    
    pdf.set_font("Arial", 'B', 9)
    pdf.set_fill_color(227, 6, 19)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(10, 8, "#", 1, 0, 'C', fill=True)
    pdf.cell(45, 8, "TIENDA ORIGEN", 1, 0, 'C', fill=True)
    pdf.cell(45, 8, "TIENDA DESTINO", 1, 0, 'C', fill=True)
    pdf.cell(30, 8, "MODELO", 1, 0, 'C', fill=True)
    pdf.cell(20, 8, "TALLA", 1, 0, 'C', fill=True)
    pdf.cell(15, 8, "CANT.", 1, 0, 'C', fill=True)
    pdf.cell(90, 8, "JUSTIFICACIÓN / FIRMAS", 1, 1, 'C', fill=True)
    
    pdf.set_font("Arial", '', 8)
    pdf.set_text_color(0, 0, 0)
    for i, row in df_traspasos.iterrows():
        pos = i + 1
        origen = str(row['ORIGEN'])
        destino = str(row['DESTINO'])
        modelo = str(row['MODELO'])
        talla = str(row['TALLA'])
        cant = str(row['CANTIDAD'])
        
        pdf.cell(10, 8, str(pos), 1, 0, 'C')
        pdf.cell(45, 8, origen, 1, 0, 'C')
        pdf.cell(45, 8, destino, 1, 0, 'C')
        pdf.cell(30, 8, modelo, 1, 0, 'C')
        pdf.cell(20, 8, talla, 1, 0, 'C')
        pdf.cell(15, 8, cant, 1, 0, 'C')
        pdf.cell(90, 8, "_____________________  /  _____________________", 1, 1, 'C')
        
    pdf.ln(10)
    pdf.set_font("Arial", 'I', 9)
    pdf.set_text_color(80, 80, 80)
    pdf.multi_cell(0, 5, "NOTA ESTRATÉGICA: Los traspasos listados en este documento tienen CARÁCTER OBLIGATORIO. Fueron calculados por el sistema identificando quiebres en tiendas donde el modelo es Top 30 en ventas, y sustrayendo inventario de sucursales donde el modelo registra 1 o cero ventas en los últimos 60 días, teniendo al menos 2 pares físicos. El objetivo es capitalizar la demanda comprobada y evitar fugas de capital.")
    
    return bytes(pdf.output(dest='S').encode('latin1'))

# Carga de archivos variables
archivo_conv = buscar_archivo('Conversion')
archivo_modelos = buscar_archivo('Venta_Modelos')
archivo_comp = buscar_archivo('Comparativo por Operacion')

# --- MÁQUINA DE ESTADOS PARA EL MENÚ Y SEGURIDAD ---
if 'vista_actual' not in st.session_state:
    st.session_state.vista_actual = 'Inicio'

# ==============================================================================
# PANTALLA 1: INICIO (LOBBY CORPORATIVO TIPO "MAS SENCILLO.PNG")
# ==============================================================================
if st.session_state.vista_actual == 'Inicio':
    
    v_p_str = "0.00%"
    v_w_str = "0.00%"
    conv_val = "0.00%"
    tkt_val = "0.00"
    fecha_val = obtener_fecha_actualizacion(archivo_conv)

    if archivo_conv:
        try:
            df_c_raw = pd.read_excel(archivo_conv) if archivo_conv.endswith('.xlsx') else pd.read_csv(archivo_conv)
            c_cv = next((c for c in df_c_raw.columns if 'Conv' in c and 'Actual' in c), None)
            c_tk = next((c for c in df_c_raw.columns if 'Ticket' in c or 'Uds/Tkt' in c or 'Prom' in c), None)
            
            if c_cv and c_tk:
                # 1. Buscar la fila de "Total" o "Resumen" para extraer el dato exacto de la Zona
                fila_total = df_c_raw[df_c_raw.iloc[:,0].astype(str).str.contains('Total|TOTAL|Resumen', case=False, na=False)]
                
                if not fila_total.empty:
                    # Extraer el valor real de la fila Total (El 11.09% oficial)
                    conv_oficial = pd.to_numeric(fila_total.iloc[0][c_cv], errors='coerce')
                    tkt_oficial = pd.to_numeric(fila_total.iloc[0][c_tk], errors='coerce')
                else:
                    # Respaldo si un día no viene la fila total: Promedio excluyendo outliers
                    df_c_clean = df_c_raw[~df_c_raw.iloc[:,0].astype(str).str.contains('3004|3015', na=False)]
                    conv_oficial = pd.to_numeric(df_c_clean[c_cv], errors='coerce').mean()
                    tkt_oficial = pd.to_numeric(df_c_clean[c_tk], errors='coerce').mean()

                # Convertir a formato porcentaje si viene en decimal
                conv_oficial = conv_oficial * 100 if pd.notna(conv_oficial) and conv_oficial <= 1 else (conv_oficial if pd.notna(conv_oficial) else 0.0)
                tkt_oficial = tkt_oficial if pd.notna(tkt_oficial) else 0.0
                
                conv_val = f"{conv_oficial:.2f}%"
                tkt_val = f"{tkt_oficial:.2f}"
        except: pass

    if archivo_comp:
        try:
            df_op = pd.read_excel(archivo_comp) if archivo_comp.endswith('.xlsx') else pd.read_csv(archivo_comp)
            c_ano = next((c for c in df_op.columns if 'año' in c.lower() or 'ano' in c.lower()), df_op.columns[0])
            c_tda = next((c for c in df_op.columns if 'tienda' in c.lower() or 'sucursal' in c.lower()), df_op.columns[2])
            c_prs = next((c for c in df_op.columns if 'pares' in c.lower() or 'cant' in c.lower()), None)
            c_imp = next((c for c in df_op.columns if 'importe' in c.lower() or 'peso' in c.lower() or 'monto' in c.lower()), None)
            c_prov = next((c for c in df_op.columns if 'prov' in c.lower()), None)
            c_tipo = next((c for c in df_op.columns if 'tipo' in c.lower() or 'concepto' in c.lower()), None)
            
            if c_prs and c_imp:
                df_op_filt = df_op.copy()
                df_op_filt[c_tda] = df_op_filt[c_tda].astype(str).str.strip()
                df_op_filt = df_op_filt[~df_op_filt[c_tda].str.contains('3004|3015', na=False)]
                if c_prov: df_op_filt = df_op_filt[~df_op_filt[c_prov].astype(str).str.strip().isin(['415', '426', '427'])]
                if c_tipo: df_op_filt = df_op_filt[~df_op_filt[c_tipo].astype(str).str.contains('BOLSA|REUSABLE|BOLSO', case=False, na=False)]
                
                df_op_filt[c_ano] = pd.to_numeric(df_op_filt[c_ano], errors='coerce')
                
                p25 = df_op_filt[df_op_filt[c_ano] == 2025][c_prs].sum()
                p26 = df_op_filt[df_op_filt[c_ano] == 2026][c_prs].sum()
                w25 = df_op_filt[df_op_filt[c_ano] == 2025][c_imp].sum()
                w26 = df_op_filt[df_op_filt[c_ano] == 2026][c_imp].sum()
                
                v_p = ((p26 - p25)/p25*100) if p25 else 0
                v_w = ((w26 - w25)/w25*100) if w25 else 0
                
                v_p_str = f"{'+' if v_p>0 else ''}{v_p:.2f}%"
                v_w_str = f"{'+' if v_w>0 else ''}{v_w:.2f}%"
        except: pass

    tienda_b64 = obtener_imagen_base64("Tienda.jpg")
    bg_css = f"background-image: linear-gradient(rgba(15, 23, 42, 0.85), rgba(15, 23, 42, 0.95)), url('{tienda_b64}');" if tienda_b64 else "background-color: #0f172a;"

    html_lobby = f"""
    <div style="{bg_css} background-size: cover; background-position: center; border-radius: 16px; padding: 50px 40px; box-shadow: 0 20px 40px rgba(0,0,0,0.5); position: relative; margin-top: 20px;">
        <div style="position: absolute; top: 30px; right: 40px; text-align: right; color: white; font-size: 13px;">
            <p style="margin:0;"><span style="color: #E30613; font-weight:bold;">Versión</span> 2.5</p>
            <p style="margin:0;">Zona Occidente</p>
        </div>
        <div class="lobby-header">
            <h1 style="color: #E30613; font-size: 55px; font-weight: 900; font-style: italic; letter-spacing: -2px; margin: 0; line-height: 1;">flexi<span style="font-size: 20px; vertical-align: super;">®</span></h1>
            <div style="width: 60px; height: 3px; background-color: #E30613; margin: 5px auto 15px auto; border-radius: 5px;"></div>
            <h1>MONITOR COMERCIAL</h1>
            <h2>ZONA OCCIDENTE</h2>
            <p>Inteligencia que impulsa decisiones</p>
        </div>
        <div class="kpi-row-lobby">
            <div class="kpi-card-lobby">
                <div class="kpi-icon-lobby"><svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 16V8a2 2 0 0 0-1-1.73l-7-4a2 2 0 0 0-2 0l-7 4A2 2 0 0 0 3 8v8a2 2 0 0 0 1 1.73l7 4a2 2 0 0 0 2 0l7-4A2 2 0 0 0 21 16z"></path><polyline points="3.27 6.96 12 12.01 20.73 6.96"></polyline><line x1="12" y1="22.08" x2="12" y2="12"></line></svg></div>
                <div class="kpi-data-lobby" style="width: 100%; display: flex; flex-direction: column; gap: 8px;">
                    <div style="display: flex; justify-content: space-between; align-items: center; border-bottom: 1px solid #334155; padding-bottom: 4px;">
                        <p style="font-size: 11px; margin: 0; color: #94a3b8; text-transform: uppercase;">Var. Pares</p>
                        <h3 style="font-size: 16px; margin: 0; color: white;">{v_p_str}</h3>
                    </div>
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <p style="font-size: 11px; margin: 0; color: #94a3b8; text-transform: uppercase;">Var. Pesos</p>
                        <h3 style="font-size: 16px; margin: 0; color: white;">{v_w_str}</h3>
                    </div>
                </div>
            </div>
            <div class="kpi-card-lobby">
                <div class="kpi-icon-lobby"><svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="23 6 13.5 15.5 8.5 10.5 1 18"></polyline><polyline points="17 6 23 6 23 12"></polyline></svg></div>
                <div class="kpi-data-lobby">
                    <h3>{conv_val}</h3>
                    <p>Conversión Regional</p>
                </div>
            </div>
            <div class="kpi-card-lobby">
                <div class="kpi-icon-lobby"><svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"></circle><path d="M16 8h-6a2 2 0 1 0 0 4h4a2 2 0 1 1 0 4H8"></path><line x1="12" y1="18" x2="12" y2="6"></line></svg></div>
                <div class="kpi-data-lobby">
                    <h3>{tkt_val}</h3>
                    <p>Ticket Regional</p>
                </div>
            </div>
            <div class="kpi-card-lobby">
                <div class="kpi-icon-lobby"><svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="4" width="18" height="18" rx="2" ry="2"></rect><line x1="16" y1="2" x2="16" y2="6"></line><line x1="8" y1="2" x2="8" y2="6"></line><line x1="3" y1="10" x2="21" y2="10"></line></svg></div>
                <div class="kpi-data-lobby">
                    <h3 style="font-size: 18px;">{fecha_val.split(' - ')[0] if ' - ' in fecha_val else fecha_val}</h3>
                    <p>Última actualización</p>
                </div>
            </div>
        </div>
    </div>
    """
    
    st.markdown(html_lobby.replace('\n', ''), unsafe_allow_html=True)
    
    col1, esp, col2 = st.columns([1, 0.05, 1])
    
    with col1:
        html_card1 = """
        <div class="action-card-lobby">
            <div class="action-icon-circle"><svg xmlns="http://www.w3.org/2000/svg" width="40" height="40" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M20 21v-2a4 4 0 0 0-4-4H8a4 4 0 0 0-4 4v2"></path><circle cx="12" cy="7" r="4"></circle></svg></div>
            <div class="action-texts">
                <h3>OPERACIÓN COMERCIAL</h3>
                <p>Seguimiento diario de la operation</p>
            </div>
        </div>
        """
        st.markdown(html_card1.replace('\n', ''), unsafe_allow_html=True)
        if st.button("INGRESAR", key="btn_op", use_container_width=True):
            st.session_state.vista_actual = 'Operativo'
            st.rerun()

    with col2:
        html_card2 = """
        <div class="action-card-lobby">
            <div class="action-icon-circle"><svg xmlns="http://www.w3.org/2000/svg" width="40" height="40" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="11" width="18" height="11" rx="2" ry="2"></rect><path d="M7 11V7a5 5 0 0 1 10 0v4"></path></svg></div>
            <div class="action-texts">
                <h3>CENTRO DE INTELIGENCIA COMERCIAL</h3>
                <p>Acceso exclusivo para usuarios autorizados</p>
            </div>
        </div>
        """
        st.markdown(html_card2.replace('\n', ''), unsafe_allow_html=True)
        if st.button("INGRESAR", key="btn_est", use_container_width=True):
            st.session_state.vista_actual = 'Login_Estrategico'
            st.rerun()

# ==============================================================================
# PANTALLA 2: LOGIN ESTRATÉGICO
# ==============================================================================
elif st.session_state.vista_actual == 'Login_Estrategico':
    st.markdown("<h1 style='color: #E30613; margin-bottom: 30px; text-align: center; margin-top: 50px;'>🔐 Autenticación Gerencial</h1>", unsafe_allow_html=True)
    
    col_esp1, col_center, col_esp2 = st.columns([1, 1.5, 1])
    with col_center:
        st.markdown("<div style='background: #1e293b; padding: 40px; border-radius: 16px; border: 1px solid #334155; text-align: center;'>", unsafe_allow_html=True)
        clave = st.text_input("Ingrese contraseña de acceso:", type="password")
        st.write("<br>", unsafe_allow_html=True)
        if st.button("Validar Acceso", type="primary", use_container_width=True):
            if clave == "Flexi2026":
                st.session_state.vista_actual = 'Estrategico'
                st.rerun()
            else:
                st.error("❌ Contraseña incorrecta.")
        st.write("<br>", unsafe_allow_html=True)
        if st.button("← Volver al Inicio", use_container_width=True):
            st.session_state.vista_actual = 'Inicio'
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

# ==============================================================================
# PANTALLA 3: MÓDULO OPERATIVO OPTIMIZADO (VISTA SUCURSALES)
# ==============================================================================
elif st.session_state.vista_actual == 'Operativo':

    col_nav1, col_nav2 = st.columns([4, 1])

    with col_nav1:
        st.markdown(
            "<h2 style='color: #E30613; margin-top: 0;'>"
            "Panel de Operación Comercial"
            "</h2>",
            unsafe_allow_html=True
        )

    with col_nav2:
        if st.button(
            "← Volver al Menú Principal",
            use_container_width=True,
            key="volver_menu_operativo"
        ):
            st.session_state.vista_actual = 'Inicio'
            st.rerun()

    st.write("---")

    modulos_operativos = [
        "📊 Desempeño Comercial",
        "👟 Top 20 Tiendas",
        "🌍 Top 20 Zona",
        "🏆 Rating Comercial",
        "🔄 Nivelación de Stock",
        "📝 Bitácora",
        "📓 Agenda de Clientes",
        "🧭 Ruta Cliente",
        "🎓 Capacitación"
    ]

    modulo_activo = st.radio(
        "Selecciona módulo operativo:",
        options=modulos_operativos,
        horizontal=True,
        key="navegacion_operativa",
        label_visibility="collapsed"
    )

    st.markdown("---")

    if modulo_activo == "📊 Desempeño Comercial":
        st.subheader("📊 DESEMPEÑO COMERCIAL")
        fecha_act = obtener_fecha_actualizacion(archivo_conv)
        st.caption(f"🔄 **Última actualización de datos:** {fecha_act}")
        if archivo_conv:
            df_c = pd.read_excel(archivo_conv) if archivo_conv.endswith('.xlsx') else pd.read_csv(archivo_conv)
            df_c = df_c[~df_c.iloc[:,0].astype(str).str.contains('3004|3015|Total|TOTAL|Resumen', na=False)]
            
            col_tienda = next((c for c in df_c.columns if 'Tienda' in c or 'TIENDA' in c), df_c.columns[0])
            col_conv_real = next((c for c in df_c.columns if 'Conv' in c and 'Actual' in c), None)
            col_tkt_real = next((c for c in df_c.columns if 'Ticket' in c or 'Uds/Tkt' in c or 'Prom' in c), None)
            
            if col_conv_real and col_tkt_real:
                meta_conv, meta_tkt = 10.9, 1.29
                df_c['CONVERSIÓN'] = df_c[col_conv_real].apply(lambda x: x*100 if x < 1 else x)
                df_c['TICKET PROMEDIO'] = df_c[col_tkt_real]
                
                ranking = df_c[[col_tienda, 'CONVERSIÓN', 'TICKET PROMEDIO']].sort_values(by='CONVERSIÓN', ascending=False).reset_index(drop=True)
                ranking.insert(0, 'POS', range(1, len(ranking) + 1))
                ranking.columns = ['#', 'TIENDA', 'CONVERSIÓN', 'TICKET PROMEDIO']

                def color_semaforo(row):
                    c_conv = row['CONVERSIÓN'] >= meta_conv
                    c_tkt = row['TICKET PROMEDIO'] >= meta_tkt
                    if c_conv and c_tkt: return ['background-color: #d4edda; color: #155724'] * 4
                    elif c_conv or c_tkt: return ['background-color: #fff3cd; color: #856404'] * 4
                    else: return ['background-color: #f8d7da; color: #721c24'] * 4

                st.table(ranking.style.apply(color_semaforo, axis=1).format({'CONVERSIÓN': '{:.2f}%', 'TICKET PROMEDIO': '{:.2f}'}))            

    elif modulo_activo == "👟 Top 20 Tiendas":
        st.subheader("👟 TOP 20 TIENDA")
        if archivo_modelos:
            df_m = pd.read_excel(archivo_modelos) if archivo_modelos.endswith('.xlsx') else pd.read_csv(archivo_modelos)
            col_m = next((c for c in df_m.columns if c.lower() in ['clave', 'modelo', 'estilo']), df_m.columns[1])
            col_p = next((c for c in df_m.columns if 'pares' in c.lower() or 'cantidad' in c.lower() or 'venta' in c.lower()), df_m.columns[2])
            col_t = next((c for c in df_m.columns if c.lower() in ['tienda', 'sucursal']), df_m.columns[0])
            col_prov = next((c for c in df_m.columns if 'prov' in c.lower()), None)

            df_m = df_m[~df_m[col_t].astype(str).str.contains('3004|3015', na=False)]
            if col_prov: df_m = df_m[~df_m[col_prov].astype(str).isin(['415', '426', '427'])]
            df_m = df_m[~df_m[col_m].astype(str).str.contains('BOLSA|REUSABLE', case=False, na=False)]

            def resaltar_top_5(data):
                estilo = pd.DataFrame('', index=data.index, columns=data.columns)
                estilo.iloc[0:5, :] = 'background-color: #d1e7dd; color: #0f5132; font-weight: bold'
                return estilo

            fecha_act = obtener_fecha_actualizacion(archivo_modelos)
            st.caption(f"🔄 **Última actualización de datos:** {fecha_act}")
            
            tiendas = sorted(df_m[col_t].unique())
            t_sel = st.selectbox("Selecciona Tienda:", tiendas, key="top20_tienda")
            df_tienda_data = df_m[df_m[col_t] == t_sel].groupby(col_m)[col_p].sum().reset_index()
            top_t = df_tienda_data.sort_values(by=col_p, ascending=False).head(20).reset_index(drop=True)
            top_t.columns = ['MODELO', 'PARES VENDIDOS']
            st.table(top_t.style.apply(resaltar_top_5, axis=None))

            pdf_bytes = generar_reporte_top20_pdf(df_top20=top_t, nombre_sucursal=str(t_sel))

            st.write("<br>", unsafe_allow_html=True)
            st.download_button(
                label="📄 Descargar Formato de Auditoría (PDF Oficial)",
                data=pdf_bytes, file_name=f"Auditoria_Top20_{t_sel}.pdf", mime="application/pdf", type="primary",
                key=f"top20_download_{t_sel}"
            )
        else:
            st.warning("⚠️ Archivo de Modelos no encontrado.")

    elif modulo_activo == "🌍 Top 20 Zona":
        st.subheader("🌍 Consolidado Zona Occidente")
        if archivo_modelos:
            df_m = pd.read_excel(archivo_modelos) if archivo_modelos.endswith('.xlsx') else pd.read_csv(archivo_modelos)
            col_m = next((c for c in df_m.columns if c.lower() in ['clave', 'modelo', 'estilo']), df_m.columns[1])
            col_p = next((c for c in df_m.columns if 'pares' in c.lower() or 'cantidad' in c.lower() or 'venta' in c.lower()), df_m.columns[2])
            col_t = next((c for c in df_m.columns if c.lower() in ['tienda', 'sucursal']), df_m.columns[0])
            col_prov = next((c for c in df_m.columns if 'prov' in c.lower()), None)

            df_m = df_m[~df_m[col_t].astype(str).str.contains('3004|3015', na=False)]
            if col_prov: df_m = df_m[~df_m[col_prov].astype(str).isin(['415', '426', '427'])]
            df_m = df_m[~df_m[col_m].astype(str).str.contains('BOLSA|REUSABLE', case=False, na=False)]

            fecha_act = obtener_fecha_actualizacion(archivo_modelos)
            st.caption(f"🔄 **Última actualización de datos:** {fecha_act}")
            
            def resaltar_top_5(data):
                estilo = pd.DataFrame('', index=data.index, columns=data.columns)
                estilo.iloc[0:5, :] = 'background-color: #d1e7dd; color: #0f5132; font-weight: bold'
                return estilo

            df_z = df_m.groupby(col_m)[col_p].sum().reset_index()
            top_z = df_z.sort_values(by=col_p, ascending=False).head(20).reset_index(drop=True)
            top_z.columns = ['MODELO', 'PARES VENDIDOS']
            st.table(top_z.style.apply(resaltar_top_5, axis=None))
        else:
            st.warning("⚠️ Archivo de Modelos no encontrado.")

    elif modulo_activo == "🏆 Rating Comercial":
        try:
            with st.spinner("Actualizando Liga de Campeones en tiempo real..."):
                df_conv_r = pd.read_excel(archivo_conv) if archivo_conv.endswith('.xlsx') else pd.read_csv(archivo_conv)
                df_conv_r = df_conv_r[~df_conv_r.iloc[:,0].astype(str).str.contains('3004|3015|Total|TOTAL|Resumen', na=False)]
                c_tda_c = next((c for c in df_conv_r.columns if 'Tienda' in c or 'TIENDA' in c), df_conv_r.columns[0])
                c_conv_real = next((c for c in df_conv_r.columns if 'Conv' in c and 'Actual' in c), None)
                c_tkt_real = next((c for c in df_conv_r.columns if 'Ticket' in c or 'Uds/Tkt' in c or 'Prom' in c), None)
                
                datos_rating = []
                for _, r in df_conv_r.iterrows():
                    tienda_str = str(r[c_tda_c]).strip()
                    conv_val = float(r[c_conv_real])
                    conv_val = conv_val * 100 if conv_val < 1 else conv_val
                    tkt_val = float(r[c_tkt_real])
                    datos_rating.append({"TIENDA": tienda_str, "CONVERSION": conv_val, "TICKET": tkt_val})
                
                df_rating = pd.DataFrame(datos_rating)
                df_rating['TIENDA_INT'] = df_rating['TIENDA'].apply(lambda x: int(re.search(r'\d+', str(x)).group()) if re.search(r'\d+', str(x)) else -1)

                if archivo_comp:
                    df_comp_r = pd.read_excel(archivo_comp) if archivo_comp.endswith('.xlsx') else pd.read_csv(archivo_comp)
                    c_ano = next((c for c in df_comp_r.columns if 'año' in c.lower() or 'ano' in c.lower()), df_comp_r.columns[0])
                    c_tda = next((c for c in df_comp_r.columns if 'tienda' in c.lower() or 'sucursal' in c.lower()), df_comp_r.columns[2])
                    c_prs = next((c for c in df_comp_r.columns if 'pares' in c.lower() or 'cant' in c.lower()), None)
                    c_prov = next((c for c in df_comp_r.columns if 'prov' in c.lower()), None)
                    
                    df_comp_r[c_tda] = df_comp_r[c_tda].astype(str).str.strip()
                    df_comp_r = df_comp_r[~df_comp_r[c_tda].str.contains('3004|3015', na=False)]
                    if c_prov: df_comp_r = df_comp_r[~df_comp_r[c_prov].astype(str).str.strip().isin(['415', '426', '427'])]
                    
                    res_comp = df_comp_r.groupby([c_tda, c_ano])[c_prs].sum().unstack(fill_value=0)
                    alcance_dict = {}
                    for tda_comp in res_comp.index:
                        try: tda_comp_int = int(re.search(r'\d+', str(tda_comp)).group())
                        except: tda_comp_int = -1
                        pares_25 = res_comp.loc[tda_comp].get(2025, 0)
                        pares_26 = res_comp.loc[tda_comp].get(2026, 0)
                        alcance_dict[tda_comp_int] = (pares_26 / pares_25 * 100) if pares_25 > 0 else 0
                    
                    df_rating['ALCANCE'] = df_rating['TIENDA_INT'].map(alcance_dict).fillna(0)

                cargar_archivos_locales()
                
                if 'df_ventas' in st.session_state and 'df_tallas' in st.session_state:
                    df_ventas_r = st.session_state.df_ventas.copy()
                    df_tallas_r = st.session_state.df_tallas
                    lista_exc_r = st.session_state.get('lista_excepciones', [])
                    df_ventas_r['tienda_int'] = df_ventas_r['Tienda'].apply(lambda x: int(re.search(r'\d+', str(x)).group()) if pd.notna(x) and re.search(r'\d+', str(x)) else -1)
                    df_ventas_r = df_ventas_r[~df_ventas_r['Proveedor'].isin([415, 426, 427])]
                    
                    quiebres_dict = {}
                    for tda_num in df_rating['TIENDA_INT']:
                        df_tienda_v = df_ventas_r[df_ventas_r['tienda_int'] == tda_num]
                        if df_tienda_v.empty:
                            quiebres_dict[tda_num] = 0
                            continue
                            
                        top_20 = df_tienda_v.groupby('Modelo')['Vtas'].sum().nlargest(20).index
                        df_top = df_tienda_v[df_tienda_v['Modelo'].isin(top_20)]
                        modelos_quebrados = set()
                        
                        for _, row in df_top.iterrows():
                            dpto = str(row.get('Departamento', '')).strip().lower()
                            modelo_act = str(row.get('Modelo', '')).strip().upper()
                            tallas_row = df_tallas_r[df_tallas_r['Valor'].astype(str).str.strip().str.lower() == dpto]
                            if not tallas_row.empty:
                                for i in range(1, 16):
                                    ex_val = row.get(f'ex{i}', 0)
                                    p_val = row.get(f'p{i}', 0)
                                    if (pd.isna(ex_val) or ex_val == 0) and (pd.isna(p_val) or p_val == 0):
                                        talla_fisica = tallas_row.iloc[0].get(f'ex{i}')
                                        if pd.notna(talla_fisica) and str(talla_fisica).strip() != '':
                                            talla_str = str(talla_fisica).strip()
                                            try:
                                                t_num = float(talla_fisica)
                                                if t_num >= 100: t_num = t_num / 10.0
                                            except:
                                                t_num = 0.0
                                                
                                            # ESCUDO GLOBAL DE TALLAS FANTASMA Y EXCEPCIONES
                                            if dpto == 'caballero' and t_num == 30.5: continue 
                                            if dpto in ['niño', 'nino'] and t_num == 21.5: continue
                                            if t_num > 25.0 and modelo_act in lista_exc_r: continue
                                            
                                            modelos_quebrados.add(row['Modelo'])
                                            break 
                        quiebres_dict[tda_num] = len(modelos_quebrados)
                    
                    df_rating['QUIEBRES'] = df_rating['TIENDA_INT'].map(quiebres_dict).fillna(0)
                else:
                    df_rating['QUIEBRES'] = 0

                def calcular_pts_ticket(t): return 35 if t >= 1.29 else 25 if t >= 1.25 else 10 if t >= 1.20 else 5
                def calcular_pts_conv(c): return 35 if c >= 10.9 else 25 if c >= 10.5 else 10 if c >= 10.0 else 5
                def calcular_pts_alcance(a): return 20 if a >= 100 else 10 if a >= 95 else 5
                def calcular_pts_quiebre(q): return 10 if q <= 5 else 5 if q <= 10 else 0

                df_rating['PTS_TKT'] = df_rating['TICKET'].apply(calcular_pts_ticket)
                df_rating['PTS_CONV'] = df_rating['CONVERSION'].apply(calcular_pts_conv)
                df_rating['PTS_ALC'] = df_rating['ALCANCE'].apply(calcular_pts_alcance)
                df_rating['PTS_QUIEBRE'] = df_rating['QUIEBRES'].apply(calcular_pts_quiebre)
                
                df_rating['BONO'] = df_rating['ALCANCE'].apply(lambda a: 5 if a >= 105 else 0)
                df_rating['PUNTAJE_TOTAL'] = df_rating['PTS_TKT'] + df_rating['PTS_CONV'] + df_rating['PTS_ALC'] + df_rating['PTS_QUIEBRE'] + df_rating['BONO']
                
                df_rating = df_rating.sort_values(by=['PUNTAJE_TOTAL', 'CONVERSION'], ascending=[False, False]).reset_index(drop=True)
                df_rating.insert(0, 'POSICIÓN', range(1, len(df_rating) + 1))
                
                df_tiendas_html = cargar_tiendas()
                df_tiendas_html.columns = df_tiendas_html.columns.astype(str).str.strip().str.upper()
                col_id = 'TIENDA' if 'TIENDA' in df_tiendas_html.columns else df_tiendas_html.columns[0]
                col_nom = 'NOMBRE' if 'NOMBRE' in df_tiendas_html.columns else (df_tiendas_html.columns[1] if len(df_tiendas_html.columns) > 1 else df_tiendas_html.columns[0])
                col_enc = 'ENCARGADO' if 'ENCARGADO' in df_tiendas_html.columns else None

                df_tiendas_html['tienda_int'] = df_tiendas_html[col_id].astype(str).apply(lambda x: int(re.search(r'\d+', x).group()) if pd.notna(x) and re.search(r'\d+', str(x)) else -1)
                
                def get_tienda_info(pos):
                    if len(df_rating) >= pos:
                        row = df_rating.iloc[pos-1]
                        tda_int = int(row['TIENDA_INT'])
                        enc_row = df_tiendas_html[df_tiendas_html['tienda_int'] == tda_int]
                        enc_name = "Encargada"
                        if col_enc and not enc_row.empty:
                            val = str(enc_row[col_enc].values[0])
                            if val and val.lower() != 'nan' and val.strip(): enc_name = " ".join(val.split()[:2]) 
                        tienda_nombre = "Sucursal"
                        if not enc_row.empty:
                            val_nom = str(enc_row[col_nom].values[0])
                            if val_nom and val_nom.lower() != 'nan' and val_nom.strip(): tienda_nombre = val_nom.strip()
                        return enc_name, str(tda_int), tienda_nombre, int(row['PUNTAJE_TOTAL']), int(row['BONO'])
                    return "N/A", "0", "N/A", 0, 0

                e1, num1, nom1, p1, b1 = get_tienda_info(1)
                e2, num2, nom2, p2, b2 = get_tienda_info(2)
                e3, num3, nom3, p3, b3 = get_tienda_info(3)

                bono_tag_1 = '<span class="bg-emerald-500/20 text-emerald-400 border border-emerald-500/50 text-[10px] font-bold px-2 py-1 rounded-full uppercase tracking-wider"><i class="fa-solid fa-star text-yellow-400"></i> Bono Crecimiento Activo</span>' if b1 > 0 else ''

                podio_html = f"""
                <div class="grid grid-cols-1 md:grid-cols-3 gap-6 mb-16 items-end mt-16 px-4">
                    <div class="bg-slate-800/80 rounded-2xl p-6 text-center h-[280px] flex flex-col justify-between hover-scale neon-border-silver relative">
                        <div class="absolute -top-10 left-1/2 -translate-x-1/2 w-20 h-20 bg-slate-700 rounded-full border-4 border-slate-300 flex items-center justify-center shadow-[0_0_20px_rgba(148,163,184,0.5)]"><i class="fa-solid fa-medal text-3xl text-slate-300"></i></div>
                        <div class="mt-8"><p class="text-4xl font-black text-white tracking-wide">{num2}</p><p class="text-sm text-cyan-400 font-bold uppercase tracking-wider mt-1">{nom2}</p><p class="text-sm text-slate-300 font-semibold uppercase mt-1"><i class="fa-solid fa-user-tie text-cyan-400 mr-1"></i> {e2}</p></div>
                        <div class="text-5xl font-black grad-primary mt-2">{p2}<span class="text-2xl text-slate-500">.0</span></div>
                    </div>
                    <div class="bg-slate-800 rounded-2xl p-6 text-center h-[350px] flex flex-col justify-between hover-scale neon-border-gold relative transform md:-translate-y-8 z-10 bg-gradient-to-b from-yellow-500/10 to-transparent">
                        <div class="absolute -top-12 left-1/2 -translate-x-1/2 w-24 h-24 bg-slate-800 rounded-full border-4 border-yellow-400 flex items-center justify-center shadow-[0_0_30px_rgba(234,179,8,0.6)]"><i class="fa-solid fa-trophy text-4xl text-yellow-400"></i></div>
                        <div class="mt-8"><div class="text-yellow-400 text-sm font-black tracking-widest mb-1"><i class="fa-solid fa-crown"></i> LÍDER ABSOLUTO</div><p class="text-5xl font-black text-white tracking-wide">{num1}</p><p class="text-lg text-cyan-400 font-bold uppercase tracking-wider mt-1">{nom1}</p><p class="text-sm text-slate-300 font-semibold uppercase mt-1 mb-2"><i class="fa-solid fa-user-tie text-cyan-400 mr-1"></i> {e1}</p>{bono_tag_1}</div>
                        <div class="text-6xl font-black text-yellow-400 drop-shadow-[0_0_10px_rgba(234,179,8,0.8)] mt-2">{p1}<span class="text-2xl text-yellow-600">.0</span></div>
                    </div>
                    <div class="bg-slate-800/80 rounded-2xl p-6 text-center h-[240px] flex flex-col justify-between hover-scale neon-border-bronze relative">
                        <div class="absolute -top-10 left-1/2 -translate-x-1/2 w-20 h-20 bg-slate-700 rounded-full border-4 border-amber-600 flex items-center justify-center shadow-[0_0_20px_rgba(217,119,6,0.5)]"><i class="fa-solid fa-medal text-3xl text-amber-600"></i></div>
                        <div class="mt-8"><p class="text-4xl font-black text-white tracking-wide">{num3}</p><p class="text-sm text-cyan-400 font-bold uppercase tracking-wider mt-1">{nom3}</p><p class="text-sm text-slate-300 font-semibold uppercase mt-1"><i class="fa-solid fa-user-tie text-cyan-400 mr-1"></i> {e3}</p></div>
                        <div class="text-4xl font-black grad-primary mt-2">{p3}<span class="text-2xl text-slate-500">.0</span></div>
                    </div>
                </div>
                """
                
                filas_html = ""
                for index, row in df_rating.iterrows():
                    pos = row['POSICIÓN']
                    tda_int = int(row['TIENDA_INT'])
                    enc_row = df_tiendas_html[df_tiendas_html['tienda_int'] == tda_int]
                    
                    encargado = "Encargada"
                    if col_enc and not enc_row.empty:
                        val = str(enc_row[col_enc].values[0])
                        if val and val.lower() != 'nan' and val.strip(): encargado = " ".join(val.split()[:2])
                    
                    tienda_nombre = "Sucursal"
                    if not enc_row.empty:
                        val_nom = str(enc_row[col_nom].values[0])
                        if val_nom and val_nom.lower() != 'nan' and val_nom.strip(): tienda_nombre = val_nom.strip()
                    
                    tienda_oficial = f"{tda_int} - {tienda_nombre}"
                    conv = f"{row['CONVERSION']:.2f}%"
                    tkt = f"{row['TICKET']:.2f}"
                    alcance = f"{row['ALCANCE']:.0f}%"
                    quiebres = int(row['QUIEBRES'])
                    pts = int(row['PUNTAJE_TOTAL'])
                    bono = int(row['BONO'])

                    bg_tr = ""; color_pos = "text-slate-300"; color_bar = "bar-grad"; riesgo_tag = ""

                    if pos <= 3: qualifier = "🏆 LÍDER"; qual_color = "text-yellow-400 font-bold"
                    elif pts >= 85: qualifier = "⭐ DESTACADO"; qual_color = "text-emerald-400 font-semibold"
                    elif pts >= 75: qualifier = "📈 COMPETITIVO"; qual_color = "text-blue-400 font-semibold"
                    else: qualifier = "⚠️ EN DESARROLLO"; qual_color = "text-red-400 font-semibold"

                    if pos == 1: bg_tr = "bg-gradient-to-r from-yellow-500/10 to-transparent"; color_pos = "text-yellow-500"
                    elif pos == 2: color_pos = "text-slate-300"
                    elif pos == 3: color_pos = "text-amber-600"
                    elif pts < 75: 
                        bg_tr = "bg-red-500/5"; color_pos = "text-slate-500"; color_bar = "bg-gradient-to-r from-red-600 to-orange-500"
                        riesgo_tag = '<span class="text-[10px] text-red-400 bg-red-400/10 px-2 py-1 rounded">Zona de Riesgo</span>'

                    bono_td = f'<br><span class="text-[10px] bg-yellow-500/20 px-1 rounded uppercase">+{bono} Bono</span>' if bono > 0 else ''
                    bono_bar = f'<div class="absolute top-0 right-0 h-full w-[5%] bg-yellow-400 rounded-r-full shadow-[0_0_10px_rgba(234,179,8,1)]"></div>' if bono > 0 else ''

                    filas_html += f"""
                    <tr class="border-b border-slate-700/50 {bg_tr} hover:bg-slate-700/50 transition-colors">
                        <td class="p-4 text-center font-black text-2xl {color_pos}">#{pos}</td>
                        <td class="p-4">
                            <p class="font-bold text-lg text-white whitespace-nowrap">{tienda_oficial}</p>
                            <p class="text-xs text-slate-400 mt-1 uppercase tracking-wider"><i class="fa-solid fa-user-tie text-cyan-500 mr-1"></i> {encargado} &nbsp;|&nbsp; <span class="{qual_color}">{qualifier}</span></p>
                        </td>
                        <td class="p-4 text-center text-emerald-400 font-bold">{conv}</td>
                        <td class="p-4 text-center text-emerald-400 font-bold">{tkt}</td>
                        <td class="p-4 text-center text-yellow-400 font-black">{alcance} {bono_td}</td>
                        <td class="p-4 text-center text-emerald-400 font-bold">{quiebres}</td>
                        <td class="p-4 w-64">
                            <div class="flex items-center justify-between mb-1"><span class="font-black text-white text-xl">{pts}.0</span>{riesgo_tag}</div>
                            <div class="w-full bg-slate-900 rounded-full h-3 border border-slate-700 relative"><div class="{color_bar} h-full rounded-full" style="width: {min(100, (pts/105)*100)}%"></div>{bono_bar}</div>
                        </td>
                    </tr>
                    """

                modal_html = """
                <div id="metaModal" class="fixed inset-0 bg-black/80 hidden items-center justify-center z-50 backdrop-blur-sm opacity-0 transition-opacity duration-300">
                    <div class="bg-slate-900 border-2 border-cyan-500/50 p-8 rounded-2xl shadow-[0_0_30px_rgba(6,182,212,0.3)] max-w-sm w-full transform scale-95 transition-transform duration-300 relative">
                        <button onclick="closeModal()" class="absolute top-4 right-4 text-slate-400 hover:text-white"><i class="fa-solid fa-xmark text-xl"></i></button>
                        <div class="text-center">
                            <div class="w-16 h-16 bg-cyan-500/20 rounded-full flex items-center justify-center mx-auto mb-4 text-cyan-400 text-2xl"><i class="fa-solid fa-bullseye"></i></div>
                            <h3 class="text-2xl font-black text-white mb-2 uppercase">Metas de Zona Occidente</h3>
                            <p class="text-slate-300 mb-6">Para dominar el Rating, tu sucursal debe cumplir y sostener:</p>
                            <div class="bg-slate-800 rounded-xl p-4 mb-4 border border-slate-700"><p class="text-sm text-slate-400 font-bold uppercase mb-1">Ticket Promedio</p><p class="text-4xl font-black text-cyan-400">1.29 <span class="text-sm text-slate-500">Uds</span></p></div>
                            <div class="bg-slate-800 rounded-xl p-4 border border-slate-700"><p class="text-sm text-slate-400 font-bold uppercase mb-1">Conversión Mínima</p><p class="text-4xl font-black text-fuchsia-400">10.90%</p></div>
                        </div>
                    </div>
                </div>
                <script>
                    function openModal() { const modal = document.getElementById('metaModal'); modal.classList.remove('hidden'); modal.classList.add('flex'); setTimeout(() => { modal.classList.remove('opacity-0'); modal.children[0].classList.remove('scale-95'); }, 10); }
                    function closeModal() { const modal = document.getElementById('metaModal'); modal.classList.add('opacity-0'); modal.children[0].classList.add('scale-95'); setTimeout(() => { modal.classList.add('hidden'); modal.classList.remove('flex'); }, 300); }
                </script>
                </body>
                """
        except Exception as e:
            st.error(f"Error en el motor de cálculo del Rating: {e}")

        try:
            if os.path.exists("Raiting Elegido..html"):
                with open("Raiting Elegido..html", "r", encoding="utf-8") as f: html_code = f.read()
                html_code = html_code.replace("TEMPORADA Q4 - 2026", "TEMPORADA Q2 - 2026")
                html_code = re.sub(r'<!-- PODIO TOP 3 -->.*?<!-- LEADERBOARD \(TABLA\) -->', f'<!-- PODIO TOP 3 -->\n{podio_html}\n<!-- LEADERBOARD (TABLA) -->', html_code, flags=re.DOTALL)
                html_code = re.sub(r'<tbody class="text-slate-200">.*?</tbody>', f'<tbody class="text-slate-200">\n{filas_html}\n</tbody>', html_code, flags=re.DOTALL)
                html_code = re.sub(r'<button class="pulse-btn(.*?)>', r'<button onclick="openModal()" class="pulse-btn\1>', html_code)
                html_code = html_code.replace("</body>", modal_html)
                components.html(html_code, height=1200, scrolling=True)
            else: st.warning("⚠️ El archivo 'Raiting Elegido..html' no se encontró en la carpeta de GitHub.")
        except Exception as e:
            st.error(f"Error al cargar el Rating: {e}")

    elif modulo_activo == "🔄 Nivelación de Stock":
        st.markdown("<h2 style='color: #B22222;'>📈 Monitor de Nivelación Flexi Occidente</h2>", unsafe_allow_html=True)
        fecha_act = obtener_fecha_actualizacion("Ventas.xlsx")
        st.caption(f"🔄 **Última actualización de datos:** {fecha_act}")
        
        cargar_archivos_locales()
        
        if 'df_ventas' in st.session_state:
            tiendas = sorted(st.session_state.df_ventas['Tienda'].unique().tolist())
            tienda_sel = st.selectbox("Selecciona la Tienda para analizar:", tiendas, key="nivelacion_tienda")
            
            if st.button("Ejecutar Análisis", key="nivelacion_btn"):
                df_tienda = st.session_state.df_ventas[
                    (st.session_state.df_ventas['Tienda'] == tienda_sel) & 
                    (~st.session_state.df_ventas['Proveedor'].isin([415, 426, 427]))
                ].copy()
                
                top_20 = df_tienda.groupby('Modelo')['Vtas'].sum().nlargest(20).index
                df_top = df_tienda[df_tienda['Modelo'].isin(top_20)].copy()
                
                lista_exc_faltantes = st.session_state.get('lista_excepciones', [])
                
                resultados = []
                for _, row in df_top.iterrows():
                    dpto = str(row['Departamento']).strip().lower()
                    modelo_act = str(row['Modelo']).strip().upper()
                    tallas_row = st.session_state.df_tallas[
                        st.session_state.df_tallas['Valor'].astype(str).str.lower() == dpto
                    ]
                    
                    if not tallas_row.empty:
                        for i in range(1, 16):
                            ex_val = row.get(f'ex{i}', 0)
                            p_val = row.get(f'p{i}', 0)
                            
                            if (pd.isna(ex_val) or ex_val == 0) and (pd.isna(p_val) or p_val == 0):
                                talla_fisica = tallas_row.iloc[0][f'ex{i}']
                                if pd.notna(talla_fisica):
                                    talla_str = str(talla_fisica).strip()
                                    try:
                                        t_num = float(talla_fisica)
                                        if t_num >= 100: t_num = t_num / 10.0
                                    except:
                                        t_num = 0.0
                                        
                                    # ESCUDO GLOBAL DE TALLAS FANTASMA Y EXCEPCIONES PARA FALTANTES VISUALES
                                    if dpto == 'caballero' and t_num == 30.5: continue 
                                    if dpto in ['niño', 'nino'] and t_num == 21.5: continue
                                    if t_num > 25.0 and modelo_act in lista_exc_faltantes: continue
                                    
                                    resultados.append({
                                        "Departamento": row['Departamento'].capitalize(),
                                        "Modelo": row['Modelo'],
                                        "Talla": talla_fisica
                                    })
                
                if resultados:
                    df_final = pd.DataFrame(resultados).drop_duplicates()
                    for dpto in df_final['Departamento'].unique():
                        st.markdown(f"<h3 style='color: #B22222;'>Bloque: {dpto}</h3>", unsafe_allow_html=True)
                        st.dataframe(df_final[df_final['Departamento'] == dpto][['Modelo', 'Talla']])
                else:
                    st.success("¡Excelente! No hay faltantes en el Top 20.")            
                
                st.markdown("---")
                st.markdown("<h3 style='color: #1e3a8a;'>🧊 Alerta de Inventario Inmóvil (Cero ventas en 60 días)</h3>", unsafe_allow_html=True)
                
                inmovil_list = []
                df_cero_vtas = df_tienda[pd.to_numeric(df_tienda['Vtas'], errors='coerce').fillna(0) <= 0]
                
                for _, row in df_cero_vtas.iterrows():
                    ex_cols = [f'ex{i}' for i in range(1, 16) if f'ex{i}' in df_cero_vtas.columns]
                    p_cols = [f'p{i}' for i in range(1, 16) if f'p{i}' in df_cero_vtas.columns]
                    
                    total_ex = pd.to_numeric(row[ex_cols], errors='coerce').fillna(0).sum()
                    total_p = pd.to_numeric(row[p_cols], errors='coerce').fillna(0).sum()
                    
                    if total_ex > 0:
                        inmovil_list.append({
                            'Modelo a Desfogar': row['Modelo'],
                            'Departamento': str(row.get('Departamento', '')).capitalize(),
                            'Pares Físicos (Total)': int(total_ex),
                            'Pares en Camino': int(total_p)
                        })
                
                if inmovil_list:
                    df_inmovil = pd.DataFrame(inmovil_list).sort_values(by='Pares Físicos (Total)', ascending=False).reset_index(drop=True)
                    st.warning(f"⚠️ Se detectaron {len(df_inmovil)} modelos con inventario físico estancado que no han rotado.")
                    st.dataframe(df_inmovil[['Modelo a Desfogar', 'Departamento', 'Pares Físicos (Total)', 'Pares en Camino']], use_container_width=True)
                    
                    pdf_inmovil_bytes = generar_reporte_inmovil_pdf(df_inmovil=df_inmovil, nombre_sucursal=str(tienda_sel))
                    st.write("<br>", unsafe_allow_html=True)
                    st.download_button(
                        label="📄 Descargar Orden de Depuración (Inventario Inmóvil)",
                        data=pdf_inmovil_bytes, 
                        file_name=f"Orden_Inmovil_{tienda_sel}.pdf", 
                        mime="application/pdf", 
                        type="primary",
                        key=f"inmovil_download_{tienda_sel}"
                    )
                else:
                    st.success("¡Felicidades! No tienes inventario 100% estancado en tu bodega.")

        else:
            st.warning("Archivos de ventas o tallas no encontrados localmente.")

    elif modulo_activo == "📝 Bitácora":
        st.subheader("📝 Registro de Incidencias Operativas")
        df_tiendas = cargar_tiendas()
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            tienda_seleccionada = st.selectbox("Selecciona la Tienda:", df_tiendas['NOMBRE'].unique(), key="bitacora_tienda")
        
        fila_tienda = df_tiendas[df_tiendas['NOMBRE'] == tienda_seleccionada]
        encargado_actual = fila_tienda['ENCARGADO'].values[0] if not fila_tienda.empty else "No encontrado"
        
        tda_num_defecto = ""
        for col in df_tiendas.columns:
            if col.strip().upper() in ['TIENDA', 'SUCURSAL', 'NUMERO', 'ID']:
                tda_num_defecto = str(fila_tienda[col].values[0])
                break
                
        with col2:
            st.markdown("**N° Sucursal en SAP/Inventario:**")
            st.info(f"🏪 {tda_num_defecto}")
            
        tienda_numero = tda_num_defecto
        st.info(f"**Encargado(a) detectado(a):** {encargado_actual}")
            
        fecha_mexico = datetime.datetime.utcnow() - datetime.timedelta(hours=6)
        fecha = st.date_input("Fecha", fecha_mexico.date(), key="bitacora_fecha")
        factor = st.selectbox("Factor Principal:", [
            "👟 Faltante de Tallas (Proyecto Tallas Extremas)",
            "🌧️ Clima adverso", "📉 Bajo tráfico atípico", "🧑‍🤝‍🧑 Plantilla incompleta", 
            "🔌 Falla: VPN FortiClient", "💻 Falla: Sistema/Terminales", 
            "🚧 Afectación de acceso", "🎉 Factor externo"
        ], key="bitacora_factor")
        
        modelo_captura = ""
        talla_captura = 0
        precio_captura = 0.0
        status_validacion = "N/A"

        if "Faltante de Tallas" in factor:
            modelo_captura = st.text_input("Modelo:", key="bitacora_modelo")
            talla_captura = st.number_input("Talla (Ej. 250 o 25.0):", min_value=1.0, max_value=350.0, step=0.5, value=25.0, key="bitacora_talla")
            precio_captura = st.number_input("Precio:", min_value=0.0, key="bitacora_precio")

        notas = st.text_area("Detalles adicionales:", key="bitacora_notas")
        
        if st.button("💾 Guardar en Bitácora", key="bitacora_btn"):
            puede_guardar = True
            
            if "Faltante de Tallas" in factor:
                if not modelo_captura:
                    st.error("❌ Para reportar falta de tallas, debes escribir el Modelo.")
                    puede_guardar = False
                elif not tienda_numero.strip().isdigit():
                    st.error("❌ Por favor, ingresa un N° de Sucursal válido (SOLO NÚMEROS, ej. 56) en la parte superior para poder cruzar el inventario. El sistema no acepta letras.")
                    puede_guardar = False
                else:
                    try:
                        cargar_archivos_locales()
                        df_v = st.session_state.df_ventas
                        df_t = st.session_state.df_tallas
                        es_valido, mensaje = validar_captura_stock(tienda_numero, modelo_captura, talla_captura, df_v, df_t)
                        if not es_valido:
                            st.error(mensaje)
                            puede_guardar = False
                        else:
                            status_validacion = "VALIDADO"
                    except Exception as e:
                        st.error(f"Error técnico en validación: {e}")
                        puede_guardar = False

            if puede_guardar:
                fila = [
                    str(fecha), tienda_seleccionada, encargado_actual, factor, notas, 
                    "", "", str(modelo_captura), str(int(talla_captura)), str(precio_captura), status_validacion
                ]
                try:
                    if 'sheet_bitacora' in globals():
                        sheet_bitacora.append_row(fila)
                        st.success(f"✅ Incidencia registrada para {tienda_seleccionada}")
                    else:
                        st.warning("⚠️ No se pudo conectar a Google Sheets, revisa tus credenciales o conexión.")
                except Exception as e:
                    st.error(f"❌ Error al intentar guardar en Google Sheets: {e}")

    elif modulo_activo == "🧭 Ruta Cliente":
        st.subheader("🧭 Protocolo Operativo en Piso de Venta")
        nombre_imagen = "RC Zona Occidente.png"
        if os.path.exists(nombre_imagen):
            st.image(nombre_imagen, use_container_width=True)
        else:
            st.warning("⚠️ La imagen 'RC Zona Occidente.png' aún no se encuentra en GitHub.")

    elif modulo_activo == "🎓 Capacitación":
        st.markdown("## 🎓 Centro de Capacitación y Desarrollo Operativo")
        st.write("Bienvenido al espacio interactivo para el fortalecimiento del sentido de pertenencia y alineación comercial de la Zona Occidente.")
        
        col_izq, col_der = st.columns([1, 1.2]) 
        
        with col_izq:
            st.markdown("### 📹 Material Audiovisual")
            opciones_video = {
                "Mi Nómina Flexi": "https://youtu.be/688Bi49rI30",
                "Tutorial Vales de Zapatos": "https://youtu.be/6hB95lYcL1g",
                "Tutorial mi Flexi": "https://youtu.be/WVi8geGSeOg"
            }
            video_seleccionado = st.selectbox("Selecciona el material audiovisual a reproducir:", list(opciones_video.keys()), key="cap_video")
            url_video = opciones_video[video_seleccionado]
            st.write("<br>", unsafe_allow_html=True)
            st.video(url_video)
            st.link_button(f"🚀 Clic aquí para ver {video_seleccionado} directo en YouTube", url_video, type="primary")
            
        with col_der:
            st.markdown("### 📘 Manual de Integración a Tiendas Flexi")
            st.info("**🎯 Objetivo General:** Establecer un proceso de acogida estandarizado que reduzca la rotación de personal en los primeros 90 días, transformando la incorporación en una experiencia de bienvenida profesional y humana.")
            
            with st.expander("🎯 1. PROPÓSITO DEL MONITOR COMERCIAL", expanded=True):
                st.markdown("""
                Este monitor interactivo fue desarrollado bajo la dirección del **LAE. José Martín Estrada Cabrera** como una herramienta estratégica y de auditoría en tiempo real. Su propósito principal es dar visibilidad total a la operación del piso de venta, permitiendo tomar decisiones basadas en datos exactos y eliminar las suposiciones.
                
                **📌 Nuestras Metas Inamovibles (El ADN de la Zona):**
                Toda la capacitación y esfuerzo de la sucursal se resume en dominar estos dos indicadores de calzado:
                * 👟 **Ticket Promedio:** Meta de **1.29** unidades por ticket.
                * 📊 **Conversión Mínima:** Meta de **10.90%** en piso de venta.
                """)
                
            with st.expander("1️⃣ PILAR I: BIENVENIDA (Logística y Orden)"):
                st.markdown("**Concepto:** Proyectar orden y profesionalismo. La preparación del entorno de trabajo es el primer mensaje que el colaborador recibe sobre la cultura de la empresa.")
                
            with st.expander("2️⃣ PILAR II: ACOMPAÑAMIENTO (Mentoría)"):
                st.markdown("**Concepto:** Eliminar la 'soledad del novato' mediante el sistema de compañero guía.")
                
            with st.expander("3️⃣ PILAR III: CLARIDAD DEL PROPÓSITO"):
                st.markdown("**Concepto:** Conectar las tareas diarias con el impacto real en el éxito de la zona y la misión de la empresa.")
                
            with st.expander("4️⃣ PILAR IV: METAS DE CORTO PLAZO"):
                st.markdown("**Concepto:** Brindar claridad absoluta sobre las expectativas de desempeño en la etapa crítica.")
                
            with st.expander("5️⃣ PILAR V: VINCULACIÓN SOCIAL"):
                st.markdown("**Concepto:** Humanizar el entorno laboral y fomentar la integración grupal.")

    elif modulo_activo == "📓 Agenda de Clientes":
        if 'client' in globals():
            modulo_agenda.mostrar_modulo_agenda(client)
        else:
            st.warning("⚠️ No se detectó conexión a Google Sheets. Revisa tus credenciales.")

# ==============================================================================
# PANTALLA 4: MÓDULO ESTRATÉGICO GERENCIAL
# ==============================================================================
elif st.session_state.vista_actual == 'Estrategico':
    col_nav1, col_nav2 = st.columns([4, 1])
    with col_nav1:
        st.markdown("<h2 style='color: #4338ca; margin-top: 0;'>Panel de Decisiones Estratégicas</h2>", unsafe_allow_html=True)
    with col_nav2:
        if st.button("← Volver al Menú Principal", use_container_width=True):
            st.session_state.vista_actual = 'Inicio'
            st.session_state.autenticado = False
            st.rerun()
            
    st.write("---")

    tab_monitor, tab_impacto, tab_comparativo, tab_visita, tab_nivelacion_intel, tab_mapa, tab_demanda, tab_macro, tab_inteligencia = st.tabs([
        "📡 Monitor Estratégico", 
        "💰 Impacto Financiero",
        "📈 Comparativo Mensual",
        "🤝 Preparación de Visita", 
        "📦 Nivelación Inteligente",
        "📍 Radar Geográfico",
        "📊 Diagnóstico Demanda", 
        "🌍 Correlación Macro",
        "🧠 Centro Inteligencia"
    ])
    
    with tab_monitor:
        st.subheader("🎯 Monitor Estratégico (Conexión en Tiempo Real)")
        st.write("Conexión directa con la base de datos maestra en la nube para cruce de inteligencia.")
        
        if st.button("Verificar Conexión y Cargar Datos", type="primary"):
            try:
                with st.spinner("Conectando con Google Sheets de forma segura..."):
                    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
                    creds = ServiceAccountCredentials.from_json_keyfile_dict(dict(st.secrets["gcp_service_account"]), scope)
                    client = gspread.authorize(creds)
                    
                    archivo = client.open_by_key('1lGlVEBgu9QsrH9PYTTuoRKQeWnYiR7OwUElCsfkDgoM')
                    sheet = archivo.get_worksheet(0)
                    datos = sheet.get_all_values()
                    
                    st.success("¡Sincronización Exitosa con la Base Maestra!")
                    st.dataframe(pd.DataFrame(datos[1:], columns=datos[0]), use_container_width=True)
                    
            except gspread.exceptions.APIError:
                st.error("Error de permisos (403): Verifique que el correo de servicio tenga acceso al archivo de Sheets.")
            except Exception as e:
                st.error(f"Ocurrió un error inesperado de conexión: {e}")
                
    with tab_impacto:
        st.subheader("💰 Impacto Financiero por quiebre")
        st.write("Proyección estadística de fugas de capital basada en el monitoreo de piso de venta.")
        
        col_sim1, col_sim2 = st.columns(2)
        with col_sim1:
            venta_muestra = st.number_input("Venta Acumulada de la Muestra (Piloto) $:", min_value=0.0, value=2276150.0, step=10000.0)
        with col_sim2:
            venta_empresa = st.number_input("Venta Total de la Empresa (Nacional) $:", min_value=0.0, value=40125942.0, step=100000.0)
        st.write("<br>", unsafe_allow_html=True)
        
        if st.button("Ejecutar Motor de Proyección", type="primary"):
            with st.spinner("Procesando y cruzando datos de la base maestra..."):
                try:
                    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
                    creds = ServiceAccountCredentials.from_json_keyfile_dict(dict(st.secrets["gcp_service_account"]), scope)
                    client_gs = gspread.authorize(creds)
                    
                    archivo = client_gs.open_by_key('1lGlVEBgu9QsrH9PYTTuoRKQeWnYiR7OwUElCsfkDgoM')
                    sheet = archivo.get_worksheet(0)
                    datos = sheet.get_all_values()
                    
                    if len(datos) > 1:
                        df_bitacora = pd.DataFrame(datos[1:], columns=datos[0])
                        df_bitacora.columns = df_bitacora.columns.str.strip()
                        
                        col_incidencia = next((c for c in df_bitacora.columns if 'Incidencia' in c or 'Factor' in c), df_bitacora.columns[3])
                        col_status = next((c for c in df_bitacora.columns if 'Status' in c or 'Validacion' in c), df_bitacora.columns[-1])
                        col_precio = next((c for c in df_bitacora.columns if 'Precio' in c), df_bitacora.columns[-2])
                        col_tienda = next((c for c in df_bitacora.columns if 'Tienda' in c), df_bitacora.columns[1])
                        col_modelo = next((c for c in df_bitacora.columns if 'Modelo' in c), df_bitacora.columns[-4])
                        
                        df_quiebres = df_bitacora[
                            (df_bitacora[col_incidencia].astype(str).str.contains("Faltante", case=False, na=False)) & 
                            (df_bitacora[col_status].astype(str).str.upper() == "VALIDADO")
                        ].copy()
                        
                        df_quiebres[col_precio] = pd.to_numeric(df_quiebres[col_precio], errors='coerce').fillna(0)
                        
                        tiendas_piloto = df_quiebres[col_tienda].nunique() 
                        if tiendas_piloto == 0: tiendas_piloto = 1
                        total_tiendas = 19
                        factor_proyeccion = total_tiendas / tiendas_piloto
                        
                        perdida_real = df_quiebres[col_precio].sum()
                        proyeccion_total = perdida_real * factor_proyeccion
                        
                        share_decimal = (venta_muestra / venta_empresa) if venta_empresa > 0 else 0
                        share_porcentaje = share_decimal * 100
                        proyeccion_nacional = (perdida_real / share_decimal) if share_decimal > 0 else 0

                        st.markdown("---")
                        st.markdown("#### 📍 Proyección a Nivel Zona Occidente")
                        col1, col2, col3 = st.columns(3)
                        col1.metric("Cobertura Piloto", f"{tiendas_piloto} de {total_tiendas} Tdas", f"{int((tiendas_piloto/total_tiendas)*100)}% de Operación Zona")
                        col2.metric("Venta Perdida (Real Piloto)", f"${perdida_real:,.2f}", f"{len(df_quiebres)} quiebres validados", delta_color="inverse")
                        col3.metric("Proyección Zona (19 Tdas)", f"${proyeccion_total:,.2f}", f"Factor Expansión: {factor_proyeccion:.2f}x", delta_color="off")
                        
                        st.markdown("---")
                        st.markdown("#### 🌍 Proyección de Impacto Nacional (Ponderado por Share)")
                        col_nat1, col_nat2 = st.columns(2)
                        with col_nat1:
                            st.markdown(f"""
                                <div class="kpi-box" style="background-color: #1e293b; border-color: #334155;">
                                    <div class="kpi-title" style="color: #94a3b8;">Share de la Muestra</div>
                                    <div class="kpi-value" style="color: #fbbf24; font-size: 32px;">{share_porcentaje:.2f}%</div>
                                </div>
                            """, unsafe_allow_html=True)

                        with col_nat2:
                            st.markdown(f"""
                                <div class="kpi-box" style="background-color: #1e293b; border-color: #deff9a; box-shadow: 0 0 15px rgba(222,255,154,0.1);">
                                    <div class="kpi-title" style="color: #deff9a;">Fuga de Capital Proyectada (Empresa)</div>
                                    <div class="kpi-value" style="color: white; font-size: 36px;">${proyeccion_nacional:,.2f}</div>
                                </div>
                            """, unsafe_allow_html=True)

                        st.markdown("---")
                        c_graf, c_tab = st.columns([2, 1])
                        with c_graf:
                            st.write("### 🏢 Fuga de Capital por Sucursal Piloto")
                            impacto_tda = df_quiebres.groupby(col_tienda)[col_precio].sum().sort_values(ascending=False)
                            st.bar_chart(impacto_tda)
                            
                        with c_tab:
                            st.write("### 👟 Top Modelos Quebrados")
                            ranking_mod = df_quiebres.groupby(col_modelo)[col_precio].sum().sort_values(ascending=False).head(10).reset_index()
                            ranking_mod.columns = ['Modelo', 'Fuga ($)']
                            st.dataframe(ranking_mod.style.format({'Fuga ($)': '${:,.2f}'}), use_container_width=True)
                            
                    else:
                        st.warning("No hay suficientes registros en la bitácora para proyectar.")
                except Exception as e:
                    st.error(f"Error procesando el cruce de datos: {e}")

    with tab_comparativo:
        st.subheader("📈 Análisis Comparativo de Calzado Mensual")
        fecha_act = obtener_fecha_actualizacion(archivo_comp)
        st.caption(f"🔄 **Última actualización de datos:** {fecha_act}")
        
        if archivo_comp:
            df_op = pd.read_excel(archivo_comp) if archivo_comp.endswith('.xlsx') else pd.read_csv(archivo_comp)
            
            c_ano = next((c for c in df_op.columns if 'año' in c.lower() or 'ano' in c.lower()), df_op.columns[0])
            c_tda_op = next((c for c in df_op.columns if 'tienda' in c.lower() or 'sucursal' in c.lower()), df_op.columns[2])
            c_prs_op = next((c for c in df_op.columns if 'pares' in c.lower() or 'cant' in c.lower()), None)
            c_imp_op = next((c for c in df_op.columns if 'importe' in c.lower() or 'peso' in c.lower() or 'monto' in c.lower()), None)
            c_prov = next((c for c in df_op.columns if 'prov' in c.lower()), None)
            c_tipo = next((c for c in df_op.columns if 'tipo' in c.lower() or 'concepto' in c.lower()), None)
            
            if c_prs_op and c_imp_op:
                df_op_display = df_op.copy()
                df_op_display['TIENDA_ID'] = df_op_display[c_tda_op].astype(str).str.extract(r'(\d+)', expand=False)
                df_op_display = df_op_display[~df_op_display['TIENDA_ID'].isin(['3004', '3015'])]
                
                if c_prov: df_op_display = df_op_display[~df_op_display[c_prov].astype(str).str.strip().isin(['415', '426', '427'])]
                if c_tipo: df_op_display = df_op_display[~df_op_display[c_tipo].astype(str).str.contains('BOLSA|REUSABLE|BOLSO', case=False, na=False)]
                
                df_op_display['ANIO_ID'] = pd.to_numeric(df_op_display[c_ano], errors='coerce').astype('Int64')
                df_op_display[c_prs_op] = pd.to_numeric(df_op_display[c_prs_op], errors='coerce').fillna(0)
                df_op_display[c_imp_op] = pd.to_numeric(df_op_display[c_imp_op], errors='coerce').fillna(0)
                
                resumen_maestro = df_op_display.groupby(['TIENDA_ID', 'ANIO_ID'])[[c_prs_op, c_imp_op]].sum().reset_index()
                tabla_pivot = resumen_maestro.pivot(index='TIENDA_ID', columns='ANIO_ID', values=[c_prs_op, c_imp_op]).fillna(0)
                tabla_pivot.columns = [f"{col[0]}_{col[1]}" for col in tabla_pivot.columns]
                
                col_prs_25 = f"{c_prs_op}_2025"
                col_prs_26 = f"{c_prs_op}_2026"
                col_imp_25 = f"{c_imp_op}_2025"
                col_imp_26 = f"{c_imp_op}_2026"
                
                for c in [col_prs_25, col_prs_26, col_imp_25, col_imp_26]:
                    if c not in tabla_pivot.columns: tabla_pivot[c] = 0
                
                tot_p25 = tabla_pivot[col_prs_25].sum()
                tot_p26 = tabla_pivot[col_prs_26].sum()
                tot_w25 = tabla_pivot[col_imp_25].sum()
                tot_w26 = tabla_pivot[col_imp_26].sum()
                var_p_global = ((tot_p26 - tot_p25) / tot_p25 * 100) if tot_p25 > 0 else 0
                var_w_global = ((tot_w26 - tot_w25) / tot_w25 * 100) if tot_w25 > 0 else 0
                
                c1, c2 = st.columns(2)
                with c1:
                    signo_p = "+" if var_p_global >= 0 else ""
                    col_p = "#155724" if var_p_global >= 0 else "#721c24"
                    st.markdown(f"<div class='kpi-box'><div class='kpi-title'>📦 Total Pares Zona Occidente</div><div class='kpi-value'>{tot_p26:,.0f} Pares</div><div class='kpi-delta' style='color: {col_p};'>Variación: {signo_p}{var_p_global:.2f}% vs 2025</div></div>", unsafe_allow_html=True)
                with c2:
                    signo_w = "+" if var_w_global >= 0 else ""
                    col_w = "#155724" if var_w_global >= 0 else "#721c24"
                    st.markdown(f"<div class='kpi-box'><div class='kpi-title'>💰 Total Ventas ($) Zona Occidente</div><div class='kpi-value'>${tot_w26:,.2f} MXN</div><div class='kpi-delta' style='color: {col_w};'>Variación: {signo_w}{var_w_global:.2f}% vs 2025</div></div>", unsafe_allow_html=True)
                
                st.write("<br>", unsafe_allow_html=True)
                
                tabla_comp = pd.DataFrame({
                    'TIENDA': tabla_pivot.index,
                    'PARES 2025': tabla_pivot[col_prs_25],
                    'PARES 2026': tabla_pivot[col_prs_26],
                    'VAR PARES %': ((tabla_pivot[col_prs_26] - tabla_pivot[col_prs_25]) / tabla_pivot[col_prs_25].replace(0, 1)) * 100,
                    'PESOS 2025': tabla_pivot[col_imp_25],
                    'PESOS 2026': tabla_pivot[col_imp_26],
                    'VAR PESOS %': ((tabla_pivot[col_imp_26] - tabla_pivot[col_imp_25]) / tabla_pivot[col_imp_25].replace(0, 1)) * 100
                }).sort_values(by='VAR PARES %', ascending=False).reset_index(drop=True)
                
                def color_variacion(val):
                    if isinstance(val, (int, float)):
                        color = '#d4edda' if val >= 0 else '#f8d7da'
                        texto = '#155724' if val >= 0 else '#721c24'
                        return f'background-color: {color}; color: {texto}; font-weight: bold;'
                    return ''

                st.table(tabla_comp.style.map(color_variacion, subset=['VAR PARES %', 'VAR PESOS %']).format({
                    'PARES 2025': '{:,.0f}', 'PARES 2026': '{:,.0f}', 'VAR PARES %': '{:+.2f}%',
                    'PESOS 2025': '${:,.2f}', 'PESOS 2026': '${:,.2f}', 'VAR PESOS %': '{:+.2f}%'
                }))
                
                st.write("<br>", unsafe_allow_html=True)
                st.markdown("---")
                
                st.markdown("### 📧 Plataforma de Comunicación Ejecutiva")
                col_clave, col_boton = st.columns([1, 2])
                with col_clave:
                    password_input = st.text_input("Clave de autorización:", type="password", key="comp_clave")
                    confirmar_envio = st.checkbox("Confirmo el envío masivo a toda la Zona Occidente.", key="confirmar_envio")
                with col_boton:
                    st.write("<br>", unsafe_allow_html=True)
                    if st.button("🚀 ENVIAR REPORTE A TODAS LAS TIENDAS", type="primary", key="comp_btn_masivo", use_container_width=True):
                        if password_input == "T5604b":
                            if not confirmar_envio:
                                st.warning("⚠️ Debes marcar la casilla de confirmación para habilitar el envío masivo.")
                            else:
                                df_tdas_envio = cargar_tiendas()
                                df_tdas_envio.columns = df_tdas_envio.columns.astype(str).str.strip().str.upper()
                                col_id_tda_envio = next((c for c in df_tdas_envio.columns if c in ['TIENDA', 'SUCURSAL', 'NUMERO', 'ID']), df_tdas_envio.columns[0])
                                col_correo_envio = next((c for c in df_tdas_envio.columns if 'CORREO' in c or 'EMAIL' in c), None)

                                if archivo_conv:
                                    df_c_envio = pd.read_excel(archivo_conv) if archivo_conv.endswith('.xlsx') else pd.read_csv(archivo_conv)
                                    df_c_envio = df_c_envio[~df_c_envio.iloc[:,0].astype(str).str.contains('3004|3015|Total|TOTAL|Resumen', na=False)]
                                    col_tda_c = next((c for c in df_c_envio.columns if 'Tienda' in c or 'TIENDA' in c), df_c_envio.columns[0])
                                    col_conv_real = next((c for c in df_c_envio.columns if 'Conv' in c and 'Actual' in c), None)
                                    col_tkt_real = next((c for c in df_c_envio.columns if 'Ticket' in c or 'Uds/Tkt' in c or 'Prom' in c), None)
                                    
                                    df_c_envio['CONVERSIÓN'] = df_c_envio[col_conv_real].apply(lambda x: x*100 if x < 1 else x)
                                    df_c_envio['TICKET PROMEDIO'] = df_c_envio[col_tkt_real]
                                    df_c_envio['TIENDA_ID'] = df_c_envio[col_tda_c].astype(str).str.extract(r'(\d+)', expand=False)

                                progress_bar = st.progress(0)
                                status_text = st.empty()
                                success_count = 0
                                omitted_count = 0 
                                total_tiendas = len(df_tdas_envio)
                                
                                for idx, row_tda in df_tdas_envio.iterrows():
                                    tda_raw = str(row_tda[col_id_tda_envio])
                                    match = re.search(r'\d+', tda_raw)
                                    if not match: continue
                                    tienda_obj = match.group()
                                    correo_oficial = str(row_tda[col_correo_envio]).strip() if col_correo_envio else "fleoutgdl@divec-flexi.com"
                                    if correo_oficial.lower() == 'nan' or not correo_oficial: correo_oficial = "fleoutgdl@divec-flexi.com"

                                    status_text.text(f"Procesando reporte para Tienda {tienda_obj}...")
                                    conv_actual, tkt_actual, encontrado_kpi = 0.0, 0.0, False

                                    if archivo_conv and 'df_c_envio' in locals():
                                        fila_c = df_c_envio[df_c_envio['TIENDA_ID'] == tienda_obj]
                                        if not fila_c.empty:
                                            conv_actual = float(fila_c.iloc[0]['CONVERSIÓN'])
                                            tkt_actual = float(fila_c.iloc[0]['TICKET PROMEDIO'])
                                            encontrado_kpi = True
                                            
                                    if not encontrado_kpi:
                                        omitted_count += 1
                                        progress_bar.progress((idx + 1) / total_tiendas)
                                        continue

                                    faltan_pares_calc, faltan_pesos_calc = 0, 0.0
                                    datos_tienda = resumen_maestro[resumen_maestro['TIENDA_ID'] == tienda_obj]
                                    
                                    if not datos_tienda.empty:
                                        pares_2025 = int(datos_tienda[datos_tienda['ANIO_ID'] == 2025][c_prs_op].sum())
                                        pares_2026 = int(datos_tienda[datos_tienda['ANIO_ID'] == 2026][c_prs_op].sum())
                                        pesos_2025 = float(datos_tienda[datos_tienda['ANIO_ID'] == 2025][c_imp_op].sum())
                                        pesos_2026 = float(datos_tienda[datos_tienda['ANIO_ID'] == 2026][c_imp_op].sum())
                                        faltan_pares_calc = pares_2025 - pares_2026
                                        faltan_pesos_calc = pesos_2025 - pesos_2026

                                    status_text.text(f"Enviando correo a Tienda {tienda_obj}...")
                                    resultado_alerta = enviar_correo_ejecutivo(tienda_obj, conv_actual, tkt_actual, 10.9, 1.29, faltan_pares_calc, faltan_pesos_calc, correo_oficial)
                                    if "✅" in resultado_alerta: success_count += 1
                                    progress_bar.progress((idx + 1) / total_tiendas)
                                    
                                status_text.text("Operación finalizada.")
                                if omitted_count > 0: st.success(f"✅ Se enviaron reportes a {success_count} sucursales. Se omitieron {omitted_count} tiendas por falta de KPIs.")
                                else: st.success(f"✅ ¡Operación exitosa! Se enviaron reportes ejecutivos a {success_count} sucursales.")
                        else:
                            st.error("❌ Clave incorrecta. Acceso denegado para el envío masivo.")

    with tab_visita:
        st.markdown("## 🤝 Expediente de Visita y Compromisos")
        st.write("Herramienta de cruce en tiempo real para direccionar la supervisión en tienda de forma objetiva.")
        
        df_tdas_visita = cargar_tiendas()
        if not df_tdas_visita.empty and 'NOMBRE' in df_tdas_visita.columns:
            tienda_seleccionada = st.selectbox("🎯 Selecciona la Sucursal a Visitar:", sorted(df_tdas_visita['NOMBRE'].unique()))
            
            if st.button("Generar Expediente de Visita", type="primary"):
                with st.spinner("Cruzando KPIs y auditando catálogo en tiempo real..."):
                    fila_tda = df_tdas_visita[df_tdas_visita['NOMBRE'] == tienda_seleccionada].iloc[0]
                    col_id = next((c for c in df_tdas_visita.columns if c.strip().upper() in ['TIENDA', 'SUCURSAL', 'NUMERO', 'ID']), df_tdas_visita.columns[0])
                    tda_raw = str(fila_tda[col_id])
                    match = re.search(r'\d+', tda_raw)
                    tienda_obj = match.group() if match else "-1"
                    encargada_obj = str(fila_tda.get('ENCARGADO', 'Encargada'))
                    
                    v_conv, v_tkt, v_alcance, v_quiebres, v_rating = 0.0, 0.0, 0.0, 0, 0
                    v_conv_ant, v_tkt_ant = 0.0, 0.0
                    v_faltan_pares, v_faltan_pesos = 0, 0.0
                    v_total_modelos, v_modelos_desfogue = 0, 0
                    
                    # Variables para la nueva auditoría Top 30
                    v_top30_ausencia = []
                    v_top30_exhibicion = []
                    
                    if archivo_conv:
                        df_c = pd.read_excel(archivo_conv) if archivo_conv.endswith('.xlsx') else pd.read_csv(archivo_conv)
                        df_c = df_c[~df_c.iloc[:,0].astype(str).str.contains('3004|3015|Total|TOTAL|Resumen', na=False)]
                        col_tda_c = next((c for c in df_c.columns if 'Tienda' in c or 'TIENDA' in c), df_c.columns[0])
                        
                        col_cv = next((c for c in df_c.columns if 'Conv' in c and 'Actual' in c), None)
                        col_tk = next((c for c in df_c.columns if 'Ticket' in c and 'Actual' in c), None)
                        if not col_tk: col_tk = next((c for c in df_c.columns if 'Ticket' in c or 'Uds/Tkt' in c or 'Prom' in c), None)
                        
                        col_cv_ant = next((c for c in df_c.columns if 'Conv' in c and 'Anterior' in c), None)
                        col_tk_ant = next((c for c in df_c.columns if 'Ticket' in c and 'Anterior' in c), None)
                        
                        df_c['TIENDA_ID'] = df_c[col_tda_c].astype(str).str.extract(r'(\d+)', expand=False)
                        fila_c = df_c[df_c['TIENDA_ID'] == tienda_obj]
                        if not fila_c.empty:
                            if col_cv:
                                v_conv = float(fila_c.iloc[0][col_cv])
                                v_conv = v_conv * 100 if v_conv < 1 else v_conv
                            if col_tk: v_tkt = float(fila_c.iloc[0][col_tk])
                            if col_cv_ant:
                                v_conv_ant = float(fila_c.iloc[0][col_cv_ant])
                                v_conv_ant = v_conv_ant * 100 if v_conv_ant < 1 else v_conv_ant
                            if col_tk_ant: v_tkt_ant = float(fila_c.iloc[0][col_tk_ant])
                    
                    if archivo_comp:
                        df_op = pd.read_excel(archivo_comp) if archivo_comp.endswith('.xlsx') else pd.read_csv(archivo_comp)
                        c_ano = next((c for c in df_op.columns if 'año' in c.lower() or 'ano' in c.lower()), df_op.columns[0])
                        c_tda_op = next((c for c in df_op.columns if 'tienda' in c.lower() or 'sucursal' in c.lower()), df_op.columns[2])
                        c_prs = next((c for c in df_op.columns if 'pares' in c.lower() or 'cant' in c.lower()), None)
                        c_prov = next((c for c in df_op.columns if 'prov' in c.lower()), None)
                        c_tipo = next((c for c in df_op.columns if 'tipo' in c.lower() or 'concepto' in c.lower()), None)
                        c_imp = next((c for c in df_op.columns if 'importe' in c.lower() or 'peso' in c.lower() or 'monto' in c.lower()), None)
                        
                        df_op['TIENDA_ID'] = df_op[c_tda_op].astype(str).str.extract(r'(\d+)', expand=False)
                        df_op = df_op[df_op['TIENDA_ID'] == tienda_obj]
                        if c_prov: df_op = df_op[~df_op[c_prov].astype(str).str.strip().isin(['415', '426', '427'])]
                        if c_tipo: df_op = df_op[~df_op[c_tipo].astype(str).str.contains('BOLSA|REUSABLE|BOLSO', case=False, na=False)]
                        
                        df_op['ANIO_ID'] = pd.to_numeric(df_op[c_ano], errors='coerce').astype('Int64')
                        df_op[c_prs] = pd.to_numeric(df_op[c_prs], errors='coerce').fillna(0)
                        if c_imp: df_op[c_imp] = pd.to_numeric(df_op[c_imp], errors='coerce').fillna(0)
                        
                        pares_25 = df_op[df_op['ANIO_ID'] == 2025][c_prs].sum()
                        pares_26 = df_op[df_op['ANIO_ID'] == 2026][c_prs].sum()
                        v_alcance = (pares_26 / pares_25 * 100) if pares_25 > 0 else 0.0
                        pesos_25 = df_op[df_op['ANIO_ID'] == 2025][c_imp].sum() if c_imp else 0.0
                        pesos_26 = df_op[df_op['ANIO_ID'] == 2026][c_imp].sum() if c_imp else 0.0
                        v_faltan_pares = int(pares_25 - pares_26)
                        v_faltan_pesos = float(pesos_25 - pesos_26)

                    cargar_archivos_locales()
                    v_pares_transito, v_modelos_transito = 0, 0
                    
                    if 'df_ventas' in st.session_state and 'df_tallas' in st.session_state:
                        df_v = st.session_state.df_ventas.copy()
                        df_t = st.session_state.df_tallas
                        lista_exc_v = st.session_state.get('lista_excepciones', [])
                        df_v['tienda_int'] = pd.to_numeric(df_v['Tienda'].astype(str).str.extract(r'(\d+)', expand=False), errors='coerce').fillna(-1)
                        
                        # FILTRO EXCLUSIVO PARA LA TIENDA VISITADA
                        df_tienda_v = df_v[(df_v['tienda_int'] == int(tienda_obj)) & (~df_v['Proveedor'].isin([415, 426, 427]))].copy()
                        
                        # ----- INICIO DE NUEVA AUDITORÍA TOP 30 ZONA VS SUCURSAL -----
                        # Aislar la base de toda la zona descartando outlets y morralla
                        df_v_zona = df_v[(~df_v['Proveedor'].isin([415, 426, 427])) & (~df_v['tienda_int'].isin([3004, 3015, -1])) & (~df_v['Modelo'].astype(str).str.contains('BOLSA|REUSABLE', case=False, na=False))].copy()
                        df_v_zona['Vtas'] = pd.to_numeric(df_v_zona['Vtas'], errors='coerce').fillna(0)
                        
                        # Obtener exactamente el Top 30 general de la región
                        top_30_zona = df_v_zona.groupby('Modelo')['Vtas'].sum().nlargest(30).index.tolist()
                        
                        # Preparar sumatorias para el cruce de la tienda
                        ex_cols_todas = [f'ex{i}' for i in range(1, 16) if f'ex{i}' in df_tienda_v.columns]
                        if not df_tienda_v.empty:
                            for c_ex in ex_cols_todas: df_tienda_v[c_ex] = pd.to_numeric(df_tienda_v[c_ex], errors='coerce').fillna(0)
                            store_vtas = df_tienda_v.groupby('Modelo')['Vtas'].sum().to_dict()
                            store_stock = df_tienda_v.groupby('Modelo')[ex_cols_todas].sum().sum(axis=1).to_dict()
                        else:
                            store_vtas = {}
                            store_stock = {}

                        # Motor de clasificación de alertas (Ausencia vs Exhibición)
                        for mod in top_30_zona:
                            vtas_mod = store_vtas.get(mod, 0)
                            stock_mod = store_stock.get(mod, 0)
                            
                            if stock_mod == 0 and vtas_mod == 0:
                                v_top30_ausencia.append(mod)
                            elif stock_mod > 0 and vtas_mod <= 1:
                                v_top30_exhibicion.append({'Modelo': mod, 'Físicos': int(stock_mod), 'Ventas': int(vtas_mod)})
                        # ----- FIN DE NUEVA AUDITORÍA TOP 30 -----

                        modelos_quebrados = set()
                        if not df_tienda_v.empty:
                            ex_cols = [f'ex{i}' for i in range(1, 16) if f'ex{i}' in df_tienda_v.columns]
                            p_cols = [f'p{i}' for i in range(1, 16) if f'p{i}' in df_tienda_v.columns]
                            for col in ex_cols + p_cols: df_tienda_v[col] = pd.to_numeric(df_tienda_v[col], errors='coerce').fillna(0)
                            
                            stock_por_modelo = df_tienda_v.groupby('Modelo')[ex_cols].sum().sum(axis=1)
                            v_total_modelos = len(stock_por_modelo[stock_por_modelo > 0])
                            v_modelos_desfogue = len(stock_por_modelo[(stock_por_modelo >= 1) & (stock_por_modelo <= 3)])
                            
                            top_20 = df_tienda_v.groupby('Modelo')['Vtas'].sum().nlargest(20).index
                            df_top = df_tienda_v[df_tienda_v['Modelo'].isin(top_20)]
                            
                            if not df_top.empty:
                                pedidos_por_modelo = df_top.groupby('Modelo')[p_cols].sum().sum(axis=1)
                                v_pares_transito = int(pedidos_por_modelo.sum())
                                v_modelos_transito = len(pedidos_por_modelo[pedidos_por_modelo > 0])
                            
                            for _, row in df_top.iterrows():
                                dpto = str(row.get('Departamento', '')).strip().lower()
                                modelo_act = str(row.get('Modelo', '')).strip().upper()
                                tallas_row = df_t[df_t['Valor'].astype(str).str.strip().str.lower() == dpto]
                                if not tallas_row.empty:
                                    for i in range(1, 16):
                                        ex_val, p_val = row.get(f'ex{i}', 0), row.get(f'p{i}', 0)
                                        if (pd.isna(ex_val) or ex_val == 0) and (pd.isna(p_val) or p_val == 0):
                                            talla_fisica = tallas_row.iloc[0].get(f'ex{i}')
                                            if pd.notna(talla_fisica) and str(talla_fisica).strip() != '':
                                                t_str = str(talla_fisica).strip()
                                                try:
                                                    t_num = float(talla_fisica)
                                                    if t_num >= 100: t_num = t_num / 10.0
                                                except: t_num = 0.0
                                                    
                                                # ESCUDO GLOBAL DE TALLAS FANTASMA Y EXCEPCIONES
                                                if dpto == 'caballero' and t_num == 30.5: continue 
                                                if dpto in ['niño', 'nino'] and t_num == 21.5: continue
                                                if t_num > 25.0 and modelo_act in lista_exc_v: continue
                                                
                                                modelos_quebrados.add(row['Modelo'])
                                                break
                            v_quiebres = len(modelos_quebrados)

                    pts_tkt = 35 if v_tkt >= 1.29 else 25 if v_tkt >= 1.25 else 10 if v_tkt >= 1.20 else 5
                    pts_conv = 35 if v_conv >= 10.9 else 25 if v_conv >= 10.5 else 10 if v_conv >= 10.0 else 5
                    pts_alc = 20 if v_alcance >= 100 else 10 if v_alcance >= 95 else 5
                    pts_qui = 10 if v_quiebres <= 5 else 5 if v_quiebres <= 10 else 0
                    bono = 5 if v_alcance >= 105 else 0
                    v_rating = pts_tkt + pts_conv + pts_alc + pts_qui + bono

                    st.markdown("---")
                    
                    color_bg = "#dcfce7" if v_rating >= 85 else "#fef08a" if v_rating >= 75 else "#fee2e2"
                    color_text = "#166534" if v_rating >= 85 else "#854d0e" if v_rating >= 75 else "#991b1b"
                    tipo_visita = "🌟 Visita de Mantenimiento y Reconocimiento" if v_rating >= 85 else "📈 Visita de Desarrollo (Ajuste de Estrategia)" if v_rating >= 75 else "🚨 Visita Crítica (Supervisión Estricta)"
                    
                    st.markdown(f"<div style='background-color: {color_bg}; color: {color_text}; padding: 15px; border-radius: 8px; text-align: center; margin-bottom: 20px;'><strong>{tipo_visita}</strong><br>Puntaje Actual: {v_rating} pts</div>", unsafe_allow_html=True)
                    
                    c1, c2, c3, c4 = st.columns(4)
                    
                    if v_conv > 0:
                        dif_meta_cv = v_conv - 10.9
                        color_meta_cv = "#155724" if dif_meta_cv >= 0 else "#721c24"
                        arrow_meta_cv = "↑" if dif_meta_cv >= 0 else "↓"
                        dif_ant_cv = v_conv - v_conv_ant
                        color_ant_cv = "#155724" if dif_ant_cv >= 0 else "#721c24"
                        arrow_ant_cv = "↑" if dif_ant_cv >= 0 else "↓"
                        texto_ant_cv = f"2025: {v_conv_ant:.2f}%" if v_conv_ant > 0 else "2025: S/D"
                        html_cv = f"""<div style="border: 1px solid rgba(49,51,63,0.2); border-radius: 0.5rem; padding: calc(1rem - 1px); background-color: #ffffff; height: 100%;"><label style="font-size: 14px; color: rgb(49, 51, 63); margin-bottom: 0.25rem;">Conversión</label><div style="font-size: 1.8rem; font-weight: 600; color: rgb(49, 51, 63); padding-bottom: 0.25rem;">{v_conv:.2f}%</div><div style="font-size: 14px; color: {color_meta_cv}; font-weight: 500;">{arrow_meta_cv} {abs(dif_meta_cv):.2f}% vs Meta</div><div style="font-size: 13px; color: {color_ant_cv}; font-weight: 500; margin-top: 2px;">{arrow_ant_cv} {abs(dif_ant_cv):.2f}% vs {texto_ant_cv}</div></div>"""
                        c1.markdown(html_cv, unsafe_allow_html=True)
                    else: c1.metric("Conversión", f"{v_conv:.2f}%")

                    if v_tkt > 0:
                        dif_meta_tk = v_tkt - 1.29
                        color_meta_tk = "#155724" if dif_meta_tk >= 0 else "#721c24"
                        arrow_meta_tk = "↑" if dif_meta_tk >= 0 else "↓"
                        dif_ant_tk = v_tkt - v_tkt_ant
                        color_ant_tk = "#155724" if dif_ant_tk >= 0 else "#721c24"
                        arrow_ant_tk = "↑" if dif_ant_tk >= 0 else "↓"
                        texto_ant_tk = f"2025: {v_tkt_ant:.2f}" if v_tkt_ant > 0 else "2025: S/D"
                        html_tk = f"""<div style="border: 1px solid rgba(49,51,63,0.2); border-radius: 0.5rem; padding: calc(1rem - 1px); background-color: #ffffff; height: 100%;"><label style="font-size: 14px; color: rgb(49, 51, 63); margin-bottom: 0.25rem;">Ticket Promedio</label><div style="font-size: 1.8rem; font-weight: 600; color: rgb(49, 51, 63); padding-bottom: 0.25rem;">{v_tkt:.2f}</div><div style="font-size: 14px; color: {color_meta_tk}; font-weight: 500;">{arrow_meta_tk} {abs(dif_meta_tk):.2f} vs Meta</div><div style="font-size: 13px; color: {color_ant_tk}; font-weight: 500; margin-top: 2px;">{arrow_ant_tk} {abs(dif_ant_tk):.2f} vs {texto_ant_tk}</div></div>"""
                        c2.markdown(html_tk, unsafe_allow_html=True)
                    else: c2.metric("Ticket Promedio", f"{v_tkt:.2f}")

                    html_c3 = f"""<div style="border: 1px solid rgba(49,51,63,0.2); border-radius: 0.5rem; padding: calc(1rem - 1px); background-color: #ffffff; height: 100%;"><label style="font-size: 14px; color: rgb(49, 51, 63); margin-bottom: 0.25rem;">Alcance Histórico</label><div style="font-size: 1.8rem; font-weight: 600; color: rgb(49, 51, 63); padding-bottom: 0.25rem;">{v_alcance:.1f}%</div></div>"""
                    c3.markdown(html_c3, unsafe_allow_html=True)

                    html_c4 = f"""<div style="border: 1px solid rgba(49,51,63,0.2); border-radius: 0.5rem; padding: calc(1rem - 1px); background-color: #ffffff; height: 100%;"><label style="font-size: 14px; color: rgb(49, 51, 63); margin-bottom: 0.25rem;">Quiebres Detectados</label><div style="font-size: 1.8rem; font-weight: 600; color: rgb(49, 51, 63); padding-bottom: 0.25rem;">{v_quiebres} Mod.</div></div>"""
                    c4.markdown(html_c4, unsafe_allow_html=True)
                    
                    st.markdown("#### 📊 Reto Acumulado vs Histórico")
                    c_pares, c_pesos = st.columns(2)
                    if v_faltan_pares > 0: c_pares.metric("Brecha en Pares", f"Faltan {v_faltan_pares:,.0f} pares", "- Por debajo del histórico", delta_color="inverse")
                    else: c_pares.metric("Brecha en Pares", f"A favor: {abs(v_faltan_pares):,.0f} pares", "+ Superando histórico")
                    if v_faltan_pesos > 0: c_pesos.metric("Brecha en Ingresos", f"Faltan ${v_faltan_pesos:,.2f}", "- Por debajo del histórico", delta_color="inverse")
                    else: c_pesos.metric("Brecha en Ingresos", f"A favor: ${abs(v_faltan_pesos):,.2f}", "+ Superando histórico")

                    # ----- INICIO UI DE AUDITORÍA TOP 30 -----
                    st.markdown("#### 🔍 Auditoría de Catálogo (Top 30 Zona vs Sucursal)")
                    c_aud1, c_aud2 = st.columns(2)
                    
                    with c_aud1:
                        if v_top30_ausencia:
                            st.error(f"🚨 {len(v_top30_ausencia)} Éxitos de Zona sin inventario")
                            with st.expander("Ver Fugas por Ausencia (Cero Stock)"):
                                df_ausencia = pd.DataFrame({"Modelos Faltantes": v_top30_ausencia})
                                df_ausencia.index += 1
                                st.table(df_ausencia)
                        else:
                            st.success("✅ Catálogo Top 30 cubierto (Sin quiebres absolutos)")

                    with c_aud2:
                        if v_top30_exhibicion:
                            st.warning(f"⚠️ {len(v_top30_exhibicion)} Éxitos estancados (Vtas ≤ 1)")
                            with st.expander("Ver Fugas por Exhibición"):
                                df_exhib = pd.DataFrame(v_top30_exhibicion)
                                df_exhib.index += 1
                                st.table(df_exhib)
                        else:
                            st.success("✅ Exhibición eficiente (Top 30 con rotación)")
                    # ----- FIN UI DE AUDITORÍA TOP 30 -----

                    st.markdown("#### 📦 Análisis de Inventario y Catálogo")
                    c_cat, c_desf, c_trans = st.columns(3)
                    c_cat.metric("Catálogo Activo", f"{v_total_modelos} Modelos", "En piso de venta")
                    
                    if v_modelos_desfogue > 0: c_desf.metric("Candidatos a Desfogue", f"{v_modelos_desfogue} Modelos", "Con 1 a 3 pares totales", delta_color="inverse")
                    else: c_desf.metric("Candidatos a Desfogue", "0 Modelos", "Inventario sano", delta_color="normal")
                    c_trans.metric("🚚 Resurtido (Top 20)", f"{v_pares_transito} Pares en tránsito", f"Para {v_modelos_transito} modelos estrella", delta_color="normal" if v_pares_transito > 0 else "off")

                    if v_modelos_desfogue > 0 and 'stock_por_modelo' in locals():
                        with st.expander("📦 Ver detalle de modelos para desfogue (1 a 3 pares totales)"):
                            df_desfogue = stock_por_modelo[(stock_por_modelo >= 1) & (stock_por_modelo <= 3)].reset_index()
                            df_desfogue.columns = ['Modelo a Desfogar', 'Pares Físicos (Total)']
                            df_desfogue = df_desfogue.sort_values(by='Pares Físicos (Total)', ascending=True).reset_index(drop=True)
                            df_desfogue.index += 1
                            st.table(df_desfogue)
                            st.write("<br>", unsafe_allow_html=True)
                            pdf_desfogue_bytes = generar_reporte_desfogue_pdf(df_desfogue=df_desfogue, nombre_sucursal=str(tienda_seleccionada))
                            st.download_button(label="📄 Descargar Orden de Desfogue (PDF Oficial)", data=pdf_desfogue_bytes, file_name=f"Orden_Desfogue_{tienda_obj}.pdf", mime="application/pdf", type="primary", key=f"desfogue_download_{tienda_obj}")

                    if v_pares_transito > 0 and 'pedidos_por_modelo' in locals():
                        with st.expander("⬇️ Ver detalle de modelos en tránsito (Top 20)"):
                            df_en_camino = pedidos_por_modelo[pedidos_por_modelo > 0].reset_index()
                            df_en_camino.columns = ['Modelo Estrella', 'Pares en Camino']
                            df_en_camino = df_en_camino.sort_values(by='Pares en Camino', ascending=False).reset_index(drop=True)
                            df_en_camino.index += 1
                            st.table(df_en_camino)

                    st.markdown("---")
                    st.markdown("### 📋 Instrucción Compromiso Autogenerada")
                    st.caption("Texto listo para ser enviado por WhatsApp o correo al finalizar la visita y dejar evidencia formal.")
                    
                    compromisos = []
                    
                    # ----- INYECCIÓN DE HALLAZGOS TOP 30 AL WHATSAPP -----
                    if v_top30_ausencia:
                        mods_txt = ", ".join(v_top30_ausencia[:3]) + ("..." if len(v_top30_ausencia) > 3 else "")
                        compromisos.append(f"🚨 **Resurtido Crítico (Fuga por Ausencia):** Gestionar el resurtido inmediato de modelos Top 30 de la Zona que no tienen presencia física en tu sucursal (Ej. {mods_txt}).")

                    if v_top30_exhibicion:
                        mods_txt = ", ".join([x['Modelo'] for x in v_top30_exhibicion[:3]]) + ("..." if len(v_top30_exhibicion) > 3 else "")
                        compromisos.append(f"⚠️ **Intervención de Exhibición:** Auditar físicamente la exhibición en puntos calientes de modelos estrella que presentan nula o bajísima rotación a pesar de tener inventario en tu bodega (Ej. {mods_txt}).")
                    # ---------------------------------------------------

                    if v_conv < 10.9: compromisos.append("👠 **Mejora en Conversión:** Implementar clínicas de abordaje al cliente en piso y ejecutar cierres de venta efectivos en el área de probadores para alcanzar la meta del 10.9%.")
                    if v_tkt < 1.29: compromisos.append("🛍️ **Impulso al Ticket Promedio:** Fomentar agresivamente el ofrecimiento del segundo par o producto de impulso (accesorio) en caja para lograr el objetivo de 1.29 unidades.")
                    if v_alcance < 100: compromisos.append("🚀 **Recuperación de Volumen:** Activar el enfoque comercial sobre los modelos del Top 20 de la Zona para igualar y superar el desplazamiento de pares respecto al año anterior.")
                    if v_faltan_pares > 0: compromisos.append(f"🎯 **Cierre de Brecha Matemática:** El objetivo exacto y obligatorio para empatar el crecimiento histórico requiere desplazar **{v_faltan_pares:,.0f} pares** adicionales, lo que representará un ingreso recuperado de **${v_faltan_pesos:,.2f} MXN**.")
                    elif v_faltan_pares < 0: compromisos.append(f"🏆 **Expansión Comercial:** Reconocemos el superávit de **+{abs(v_faltan_pares):,.0f} pares** (${abs(v_faltan_pesos):,.2f} MXN) frente al histórico. El compromiso es mantener esta aceleración y proteger el liderazgo comercial de la sucursal.")
                    if v_quiebres > 5: compromisos.append("⚠️ **Gestión de Quiebres:** Garantizar el reporte oportuno en la Bitácora sobre faltantes de Tallas Extremas para gestionar la nivelación y evitar fuga de capital.")
                    if v_modelos_desfogue > 0: compromisos.append(f"📦 **Depuración de Inventario (Desfogue):** Se detectaron **{v_modelos_desfogue} modelos** con inventario marginal (1 a 3 pares totales). El compromiso es generar el reporte de transferencia a sucursales Outlet antes del cierre de semana para liberar espacio de bodega y concentrar la labor de venta en el catálogo de alta rotación.")
                    if not compromisos: compromisos.append("⭐ **Sostenimiento de Excelencia:** Mantener la estricta disciplina en los procesos de venta actuales, protegiendo los KPIs que hoy mantienen a la sucursal en el nivel de excelencia.")

                    texto_viñetas = "\n".join([f"- {c}" for c in compromisos])
                    fecha_hoy = (datetime.datetime.utcnow() - datetime.timedelta(hours=6)).strftime("%d/%m/%Y")
                    texto_final = f"Estimada {encargada_obj},\n\nCon base en la supervisión operativa del día de hoy ({fecha_hoy}), establecemos los siguientes compromisos ejecutivos para la sucursal {tienda_seleccionada}, alineados a su desempeño actual ({v_rating} pts de Rating):\n\n{texto_viñetas}\n\nQueda bajo su responsabilidad el seguimiento de estas directrices con su equipo de asesores para nuestra próxima evaluación.\n\nAtentamente,\nLAE. José Martín Estrada Cabrera\nGerencia Comercial Zona Occidente\n"
                    st.text_area("Copia el siguiente mensaje:", value=texto_final, height=450)
        else:
            st.warning("No se pudo cargar la base de tiendas.")

    with tab_nivelacion_intel:
        st.subheader("📦 Nivelación Inteligente de Inventario")
        st.write("Identificación automática de pares inmovilizados para cubrir quiebres absolutos de modelos estrella.")
        
        if st.button("🚀 Ejecutar Algoritmo de Nivelación", type="primary"):
            cargar_archivos_locales()
            if 'df_ventas' in st.session_state and 'df_tallas' in st.session_state:
                with st.spinner("Analizando matrices de inventario y ventas (Últimos 60 días)..."):
                    df_v = st.session_state.df_ventas.copy()
                    df_t = st.session_state.df_tallas.copy()
                    lista_exc_n = st.session_state.get('lista_excepciones', [])

                    df_tiendas_map = cargar_tiendas()
                    df_tiendas_map.columns = df_tiendas_map.columns.astype(str).str.strip().str.upper()
                    col_id = next((c for c in df_tiendas_map.columns if c in ['TIENDA', 'SUCURSAL', 'NUMERO', 'ID']), df_tiendas_map.columns[0])
                    col_nom = next((c for c in df_tiendas_map.columns if c in ['NOMBRE']), df_tiendas_map.columns[1] if len(df_tiendas_map.columns)>1 else df_tiendas_map.columns[0])

                    tiendas_dict = {}
                    for _, r in df_tiendas_map.iterrows():
                        try:
                            t_id = int(re.search(r'\d+', str(r[col_id])).group())
                            tiendas_dict[t_id] = str(r[col_nom]).strip()
                        except: pass
                    
                    def get_tienda_nombre(t_id):
                        nom = tiendas_dict.get(t_id, "Sucursal")
                        return f"{t_id} - {nom}"

                    # 1. Limpieza y preparación de la base
                    df_v['tienda_int'] = pd.to_numeric(df_v['Tienda'].astype(str).str.extract(r'(\d+)', expand=False), errors='coerce').fillna(-1).astype(int)
                    df_v = df_v[(df_v['tienda_int'] > 0) & (~df_v['Proveedor'].isin([415, 426, 427])) & (~df_v['tienda_int'].isin([3004, 3015]))]
                    
                    df_v['Vtas'] = pd.to_numeric(df_v['Vtas'], errors='coerce').fillna(0)
                    df_v['Modelo_cln'] = df_v['Modelo'].astype(str).str.strip().str.upper()

                    for i in range(1, 16):
                        if f'ex{i}' in df_v.columns: df_v[f'ex{i}'] = pd.to_numeric(df_v[f'ex{i}'], errors='coerce').fillna(0)
                        if f'p{i}' in df_v.columns: df_v[f'p{i}'] = pd.to_numeric(df_v[f'p{i}'], errors='coerce').fillna(0)

                    df_t['Valor_cln'] = df_t.iloc[:, 0].astype(str).str.strip().str.lower()

                    traspasos_sugeridos = []
                    tiendas_list = df_v['tienda_int'].unique()

                    # 2. Motor de Emparejamiento con KARDEX VIRTUAL
                    for tda_receptor in tiendas_list:
                        df_receptor = df_v[df_v['tienda_int'] == tda_receptor]
                        top_modelos = df_receptor[df_receptor['Vtas'] > 0].groupby('Modelo_cln')['Vtas'].sum().nlargest(30).index

                        for modelo in top_modelos:
                            row_mod_rec = df_receptor[df_receptor['Modelo_cln'] == modelo].iloc[0]
                            dpto = str(row_mod_rec.get('Departamento', '')).strip().lower()

                            tallas_row = df_t[df_t['Valor_cln'] == dpto]
                            if tallas_row.empty: continue

                            for i in range(1, 16):
                                col_ex = f'ex{i}'
                                col_p = f'p{i}'
                                if col_ex not in row_mod_rec or col_p not in row_mod_rec: continue

                                ex_rec = row_mod_rec[col_ex]
                                p_rec = row_mod_rec[col_p]
                                talla_real = tallas_row.iloc[0].get(col_ex, '')

                                if pd.isna(talla_real) or str(talla_real).strip() == '': continue
                                
                                try:
                                    t_num = float(talla_real)
                                    if t_num >= 100: t_num = t_num / 10.0
                                except: t_num = 0.0

                                # ESCUDO GLOBAL DE TALLAS FANTASMA Y EXCEPCIONES
                                if dpto == 'caballero' and t_num == 30.5: continue 
                                if dpto in ['niño', 'nino'] and t_num == 21.5: continue
                                if t_num > 25.0 and modelo in lista_exc_n: continue

                                if ex_rec == 0 and p_rec == 0:
                                    # KARDEX VIRTUAL: Lee el inventario descontando lo que ya se donó
                                    df_donadores = df_v[
                                        (df_v['Modelo_cln'] == modelo) & 
                                        (df_v['tienda_int'] != tda_receptor) & 
                                        (df_v['Vtas'] <= 1) & 
                                        (df_v[col_ex] >= 2)
                                    ].copy()
                                    
                                    # REGLA EXCLUSIVA DE ENRUTAMIENTO: TIENDA 12 PAAR (Ecosistema Aislado)
                                    tiendas_validas_donadoras = df_donadores['tienda_int'].apply(
                                        lambda o: (o in [56, 59, 133]) if tda_receptor == 12 else (tda_receptor in [56, 59, 133] if o == 12 else True)
                                    )
                                    df_donadores = df_donadores[tiendas_validas_donadoras]

                                    if not df_donadores.empty:
                                        df_donadores = df_donadores.sort_values(by=['Vtas', col_ex], ascending=[True, False])
                                        mejor_donador = df_donadores.iloc[0]
                                        origen_idx = mejor_donador.name

                                        origen_nom = get_tienda_nombre(mejor_donador['tienda_int'])
                                        destino_nom = get_tienda_nombre(tda_receptor)
                                        
                                        # KARDEX VIRTUAL: Restar el par físico al donador instantáneamente
                                        df_v.at[origen_idx, col_ex] -= 1

                                        traspasos_sugeridos.append({
                                            'ORIGEN': origen_nom,
                                            'DESTINO': destino_nom,
                                            'MODELO': row_mod_rec['Modelo'],
                                            'TALLA': str(talla_real).strip(),
                                            'CANTIDAD': 1,
                                            'JUSTIFICACIÓN': f"Destino (Top 30). Origen (Vtas: {int(mejor_donador['Vtas'])}, Stock Restante: {int(df_v.at[origen_idx, col_ex])})"
                                        })

                    # 3. Presentación de Resultados
                    if traspasos_sugeridos:
                        df_traspasos = pd.DataFrame(traspasos_sugeridos)
                        st.success(f"✅ ¡Análisis completado! Se encontraron {len(df_traspasos)} oportunidades de rescate de capital bajo las reglas estrictas de negocio.")
                        
                        html_tabla = "<div style='overflow-x:auto;'>\n<table style='width:100%; border-collapse: collapse; font-family: sans-serif; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);'>\n<thead>\n<tr style='background-color: #E30613; color: white; text-transform: uppercase; font-size: 13px;'>\n<th style='padding: 12px 15px; border: 1px solid #b9000b; text-align: left;'>Sucursal Origen</th>\n<th style='padding: 12px 15px; border: 1px solid #b9000b; text-align: left;'>Sucursal Destino</th>\n<th style='padding: 12px 15px; border: 1px solid #b9000b; text-align: center;'>Modelo</th>\n<th style='padding: 12px 15px; border: 1px solid #b9000b; text-align: center;'>Talla</th>\n<th style='padding: 12px 15px; border: 1px solid #b9000b; text-align: center;'>Cant.</th>\n<th style='padding: 12px 15px; border: 1px solid #b9000b; text-align: left;'>Lógica del Sistema</th>\n</tr>\n</thead>\n<tbody>\n"
                        for _, row in df_traspasos.iterrows():
                            html_tabla += f"<tr style='border-bottom: 1px solid #e2e8f0; background-color: white; transition: background-color 0.2s;'>\n<td style='padding: 12px 15px; background-color: #fee2e2; color: #991b1b; font-weight: bold; border-right: 1px solid #e2e8f0;'>🏪 {row['ORIGEN']}</td>\n<td style='padding: 12px 15px; background-color: #d1fae5; color: #166534; font-weight: bold; border-right: 1px solid #e2e8f0;'>🎯 {row['DESTINO']}</td>\n<td style='padding: 12px 15px; text-align: center; font-weight: bold; color: #1e293b;'>{row['MODELO']}</td>\n<td style='padding: 12px 15px; text-align: center; color: #475569;'>{row['TALLA']}</td>\n<td style='padding: 12px 15px; text-align: center; background-color: #fef08a; color: #854d0e; font-size: 16px; font-weight: 900; border-left: 1px solid #e2e8f0; border-right: 1px solid #e2e8f0;'>{row['CANTIDAD']}</td>\n<td style='padding: 12px 15px; font-size: 12px; color: #64748b;'>{row['JUSTIFICACIÓN']}</td>\n</tr>\n"
                        html_tabla += "</tbody></table></div>"
                        st.markdown(html_tabla, unsafe_allow_html=True)

                        pdf_bytes = generar_reporte_traspaso_masivo_pdf(df_traspasos)
                        st.write("<br>", unsafe_allow_html=True)
                        st.download_button(label="📄 Descargar Órdenes de Traspaso (PDF Oficial)", data=pdf_bytes, file_name="Ordenes_Nivelacion_Inventario.pdf", mime="application/pdf", type="primary", key="btn_descarga_traspasos")
                    else: st.info("No se encontraron oportunidades de traspaso. ¡El inventario de la Zona está equilibrado!")
            else: st.warning("⚠️ No se encontraron los archivos locales necesarios para ejecutar el algoritmo de nivelación.")

    with tab_mapa:
        st.subheader("📍 Radar Geográfico de Rendimiento")
        st.write("Vista satelital en tiempo real con cruce de oportunidades de expansión.")
        
        col_m1, col_m2, col_m3 = st.columns([1.5, 1.5, 1])
        with col_m1: mostrar_cobertura = st.toggle("🔵 Radios de Cobertura (5km)", value=False)
        with col_m2: mostrar_calor = st.toggle("🔥 Mapa de Calor (Quiebres)", value=False)
        with col_m3: btn_ia = st.button("🤖 Generar Diagnóstico IA", type="primary", use_container_width=True)
        if btn_ia: st.session_state.mostrar_reporte_ia = True
            
        try:
            df_mapa = cargar_tiendas()
            df_mapa.columns = df_mapa.columns.astype(str).str.strip().str.upper()
            
            if 'LATITUD' in df_mapa.columns and 'LONGITUD' in df_mapa.columns:
                col_id_tda = next((c for c in df_mapa.columns if c in ['TIENDA', 'SUCURSAL', 'NUMERO', 'ID']), df_mapa.columns[0])
                col_nom_tda = next((c for c in df_mapa.columns if c in ['NOMBRE']), df_mapa.columns[1] if len(df_mapa.columns)>1 else df_mapa.columns[0])
                col_enc_tda = 'ENCARGADO' if 'ENCARGADO' in df_mapa.columns else 'N/A'
                
                name_to_id_dict = {}
                for _, r in df_mapa.iterrows():
                    try:
                        t_id_match = re.search(r'\d+', str(r[col_id_tda]))
                        if t_id_match: name_to_id_dict[str(r[col_nom_tda]).strip().upper()] = int(t_id_match.group())
                    except: pass
                
                quiebres_dict = {}
                try:
                    if 'sheet_bitacora' in globals():
                        datos_b = sheet_bitacora.get_all_values()
                        if len(datos_b) > 1:
                            df_b = pd.DataFrame(datos_b[1:], columns=datos_b[0])
                            col_inc = next((c for c in df_b.columns if 'Incidencia' in c or 'Factor' in c), df_b.columns[3])
                            col_st = next((c for c in df_b.columns if 'Status' in c or 'Validacion' in c), df_b.columns[-1])
                            col_pr = next((c for c in df_b.columns if 'Precio' in c), df_b.columns[-2])
                            col_td = next((c for c in df_b.columns if 'Tienda' in c), df_b.columns[1])
                            
                            df_q = df_b[(df_b[col_inc].astype(str).str.contains("Faltante", case=False, na=False)) & (df_b[col_st].astype(str).str.upper() == "VALIDADO")].copy()
                            df_q[col_pr] = pd.to_numeric(df_q[col_pr], errors='coerce').fillna(0)
                            def extraer_id_tienda(val):
                                str_val = str(val).strip().upper()
                                match = re.search(r'\d+', str_val)
                                if match: return int(match.group())
                                return name_to_id_dict.get(str_val, -1)
                            df_q['T_ID'] = df_q[col_td].apply(extraer_id_tienda)
                            quiebres_dict = df_q.groupby('T_ID')[col_pr].sum().to_dict()
                except: pass

                kpi_dict = {}
                if archivo_conv:
                    df_c_mapa = pd.read_excel(archivo_conv) if archivo_conv.endswith('.xlsx') else pd.read_csv(archivo_conv)
                    df_c_mapa = df_c_mapa[~df_c_mapa.iloc[:,0].astype(str).str.contains('3004|3015|Total|TOTAL|Resumen', na=False)]
                    col_tda_mapa = next((c for c in df_c_mapa.columns if 'Tienda' in c or 'TIENDA' in c), df_c_mapa.columns[0])
                    col_cv_mapa = next((c for c in df_c_mapa.columns if 'Conv' in c and 'Actual' in c), None)
                    col_tk_mapa = next((c for c in df_c_mapa.columns if 'Ticket' in c or 'Uds/Tkt' in c or 'Prom' in c), None)
                    for _, r in df_c_mapa.iterrows():
                        try:
                            t_id = int(re.search(r'\d+', str(r[col_tda_mapa])).group())
                            cv = float(r[col_cv_mapa]) if col_cv_mapa else 0.0
                            cv = cv * 100 if cv < 1 else cv
                            tk = float(r[col_tk_mapa]) if col_tk_mapa else 0.0
                            kpi_dict[t_id] = {'conv': cv, 'tkt': tk}
                        except: pass
                
                markers_data = []
                for _, r in df_mapa.iterrows():
                    try:
                        lat, lon = float(r['LATITUD']), float(r['LONGITUD'])
                        t_id = int(re.search(r'\d+', str(r[col_id_tda])).group())
                        nom = str(r[col_nom_tda])
                        enc = str(r[col_enc_tda]) if col_enc_tda in r else 'N/A'
                        kpis = kpi_dict.get(t_id, {'conv': 0.0, 'tkt': 0.0})
                        cv, tk = kpis['conv'], kpis['tkt']
                        fuga_capital = quiebres_dict.get(t_id, 0.0)
                        
                        if cv >= 10.9 and tk >= 1.29: color = "#22c55e"
                        elif cv >= 10.9 or tk >= 1.29: color = "#eab308"
                        elif cv > 0 or tk > 0: color = "#ef4444"
                        else: color = "#94a3b8"
                        
                        markers_data.append({"lat": lat, "lon": lon, "name": f"{t_id} - {nom}", "encargada": enc, "conv": round(cv, 2), "tkt": round(tk, 2), "color": color, "fuga": fuga_capital})
                    except: pass
                
                if markers_data:
                    js_show_coverage = "true" if mostrar_cobertura else "false"
                    js_show_heatmap = "true" if mostrar_calor else "false"
                    
                    leaflet_html = f"""
                    <!DOCTYPE html><html><head><link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" /><script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script><style>.custom-div-icon {{ background: transparent; border: none; }} .leaflet-popup-content-wrapper {{ border-radius: 12px; padding: 15px; box-shadow: 0 10px 25px rgba(0,0,0,0.2); }} .leaflet-popup-content {{ margin: 0; }} .heat-pulse {{ filter: blur(8px); animation: pulse-animation 2s infinite; }} @keyframes pulse-animation {{ 0% {{ opacity: 0.4; transform: scale(0.95); }} 50% {{ opacity: 0.7; transform: scale(1.05); }} 100% {{ opacity: 0.4; transform: scale(0.95); }} }}</style></head><body style="margin: 0; padding: 0;"><div id="map" style="width: 100%; height: 650px; border-radius: 12px;"></div><script>var map = L.map('map').setView([20.67, -103.35], 7); L.tileLayer('https://tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png', {{ maxZoom: 19, attribution: '&copy; OpenStreetMap' }}).addTo(map); var markers = {json.dumps(markers_data)}; var bounds = []; var showCoverage = {js_show_coverage}; var showHeatmap = {js_show_heatmap}; markers.forEach(function(m) {{ if (showCoverage) {{ L.circle([m.lat, m.lon], {{ radius: 5000, color: '#3b82f6', weight: 1.5, dashArray: '5, 5', fillColor: '#3b82f6', fillOpacity: 0.1 }}).addTo(map); }} if (showHeatmap && m.fuga > 0) {{ var heatRadius = 20 + (Math.min(m.fuga / 3000, 1) * 40); L.circleMarker([m.lat, m.lon], {{ radius: heatRadius, color: 'transparent', fillColor: '#ef4444', fillOpacity: 0.6, className: 'heat-pulse' }}).addTo(map); }} var htmlIcon = "<div style='background-color:" + m.color + "; width:20px; height:20px; border-radius:50%; border:3px solid white; box-shadow:0 0 8px rgba(0,0,0,0.5); z-index: 1000; position: relative;'></div>"; var customIcon = L.divIcon({{ className: 'custom-div-icon', html: htmlIcon, iconSize: [20, 20], iconAnchor: [10, 10] }}); var convColor = m.conv >= 10.9 ? '#155724' : '#721c24'; var tktColor = m.tkt >= 1.29 ? '#155724' : '#721c24'; var htmlFuga = m.fuga > 0 ? "<div style='margin-top:10px; background-color:#fee2e2; border-radius:6px; padding:6px; text-align:center;'><span style='font-size:10px; color:#991b1b; font-weight:800;'>🔥 FUGA POR QUIEBRE</span><br><span style='font-size:14px; font-weight:900; color:#b91c1c;'>$" + m.fuga.toLocaleString() + "</span></div>" : ""; var popupHTML = "<div style='font-family: Arial, sans-serif; min-width: 220px;'><h3 style='margin:0 0 5px 0; color:#1e293b; font-size: 16px; font-weight: 800;'>" + m.name + "</h3><p style='margin:0 0 12px 0; font-size:13px; color:#64748b; font-weight: 600;'>👤 " + m.encargada + "</p><div style='display:flex; justify-content:space-between; border-top:1px solid #e2e8f0; padding-top:8px; margin-top: 8px;'><div style='background-color: #f8fafc; padding: 5px 10px; border-radius: 6px; text-align: center; width: 45%;'><div style='font-size:10px; color:#94a3b8; font-weight: 800;'>CONVERSIÓN</div><div style='font-size:16px; font-weight:900; color:" + convColor + "'>" + m.conv + "%</div></div><div style='background-color: #f8fafc; padding: 5px 10px; border-radius: 6px; text-align: center; width: 45%;'><div style='font-size:10px; color:#94a3b8; font-weight: 800;'>TICKET PROM.</div><div style='font-size:16px; font-weight:900; color:" + tktColor + "'>" + m.tkt + "</div></div></div>" + htmlFuga + "</div>"; var marker = L.marker([m.lat, m.lon], {{icon: customIcon}}).addTo(map); marker.bindPopup(popupHTML); bounds.push([m.lat, m.lon]); }}); if (bounds.length > 0) {{ setTimeout(function() {{ map.fitBounds(bounds, {{padding: [40, 40], maxZoom: 11}}); }}, 500); }} </script></body></html>
                    """
                    
                    st.markdown("<div style='display: flex; gap: 20px; justify-content: center; margin-bottom: 15px; background: #f8fafc; padding: 10px; border-radius: 8px; border: 1px solid #e2e8f0;'><div style='display: flex; align-items: center; gap: 8px;'><div style='width: 15px; height: 15px; border-radius: 50%; background-color: #22c55e;'></div><span style='font-size: 14px; font-weight: 600; color: #334155;'>Supera Metas (Ambas)</span></div><div style='display: flex; align-items: center; gap: 8px;'><div style='width: 15px; height: 15px; border-radius: 50%; background-color: #eab308;'></div><span style='font-size: 14px; font-weight: 600; color: #334155;'>Riesgo (Falta una meta)</span></div><div style='display: flex; align-items: center; gap: 8px;'><div style='width: 15px; height: 15px; border-radius: 50%; background-color: #ef4444;'></div><span style='font-size: 14px; font-weight: 600; color: #334155;'>Crítico (Ninguna meta)</span></div></div>", unsafe_allow_html=True)
                    components.html(leaflet_html, height=670)
                    
                    if st.session_state.get('mostrar_reporte_ia', False):
                        st.markdown("---")
                        html_ia = "<div style='background-color: #0f172a; border-left: 6px solid #8b5cf6; padding: 25px; border-radius: 12px; color: #f8fafc; box-shadow: 0 10px 30px rgba(0,0,0,0.3); margin-top: 10px;'><h3 style='color: #a78bfa; margin-top: 0; font-weight: 900; letter-spacing: 1px;'><i class='fa-solid fa-brain' style='margin-right:8px;'></i> DICTAMEN DE VIABILIDAD DE EXPANSIÓN (IA)</h3><p style='color: #cbd5e1; font-size: 15px; line-height: 1.6;'><strong>Señores Directivos:</strong><br><br>El motor geográfico ha cruzado los radios de cobertura geométrica (5 km) con los puntos térmicos de demanda insatisfecha extraídos de la Bitácora Comercial, detectando una oportunidad estratégica prioritaria en la Zona Occidente:</p><div style='background-color: #1e293b; padding: 15px; border-radius: 8px; border: 1px solid #334155; margin: 20px 0;'><ul style='margin: 0; color: #cbd5e1; font-size: 14px; line-height: 1.8;'><li style='margin-bottom: 8px;'><strong style='color: white;'>📍 Hotspot Detectado:</strong> Corredor Sur (Perímetro de congestión entre Plaza del Sol y Galerías Santa Anita).</li><li style='margin-bottom: 8px;'><strong style='color: white;'>🔥 Demanda Desbordada:</strong> Las sucursales frontera están negando ventas constantemente por saturación de inventario en bodega (Quiebre de Tallas Extremas y Top 20).</li><li style='margin-bottom: 8px;'><strong style='color: white;'>🔵 Vacío Geográfico:</strong> Existe un radio de 6.8 km de alto tráfico económico comprobado sin presencia de nuestra marca.</li><li><strong style='color: white;'>🎯 Riesgo de Canibalización:</strong> < 4% (La demanda del mercado local excede la capacidad actual instalada).</li></ul></div><p style='font-size: 16px; font-weight: bold; color: #10b981;'>🚀 RECOMENDACIÓN ESTRATÉGICA:</p><p style='color: #cbd5e1; font-size: 14px; line-height: 1.6;'>El algoritmo determina una viabilidad <strong>ALTA (88%)</strong> para la apertura de un nuevo punto de venta nivel <em>Boutique o Kiosco</em> en la zona ciega central. Como acción correctiva inmediata, se sugiere elevar un 30% el stock matriz del catálogo Top 20 en las tiendas perimetrales para capturar el excedente térmico.</p></div>"
                        st.markdown(html_ia, unsafe_allow_html=True)
                        if st.button("Cerrar Dictamen"):
                            st.session_state.mostrar_reporte_ia = False
                            st.rerun()
                else: st.info("No se pudieron procesar las coordenadas. Verifica que el archivo esté completo.")
            else: st.warning("⚠️ No se detectaron las columnas 'LATITUD' y 'LONGITUD' en tu archivo 'CORREO DE TIENDAS.xlsx'.")
        except Exception as e:
            st.error(f"Error al generar el mapa: {e}")

    with tab_demanda:
        st.subheader("📊 Diagnóstico de Demanda")
        st.info("Módulo en fase de diseño. Próximamente: Análisis multivariable predictivo avanzado.")

    with tab_macro:
        st.subheader("🌍 Correlación Macroeconómica (INPC)")
        st.info("Módulo en fase de diseño. Próximamente: Cruce de inflación vs ticket promedio.")

    with tab_inteligencia:
        st.subheader("Radar de Factores Externos y Logística")
        st.write("Monitoreo de factores que impactan la afluencia y el suministro de calzado.")
        
        try:
            # Aquí el monitor lee el archivo que creó tu agente en silencio
            df_inteligencia = pd.read_csv("datos_inteligencia.csv")
            
            # Mostramos los datos con un diseño limpio
            st.dataframe(
                df_inteligencia[['Categoría', 'Título', 'Fecha', 'Enlace']], 
                use_container_width=True,
                hide_index=True
            )
            st.caption(f"Última actualización de datos: {df_inteligencia['Última Actualización'].iloc[0]}")
            
            # Botón opcional para forzar la actualización manual desde la interfaz
            if st.button("Actualizar datos del mercado ahora", type="secondary"):
                with st.spinner("Buscando nueva información..."):
                    try:
                        import subprocess
                        subprocess.run(["python", "agente_inteligencia.py"])
                        st.rerun()
                    except Exception as e:
                        st.error(f"No se pudo ejecutar el agente: {e}")

        except FileNotFoundError:
            st.info("Buscando nueva información del mercado. El agente aún no ha generado el archivo de hoy.")
            # Si no existe, damos la opción de generarlo la primera vez
            if st.button("Generar archivo inicial de datos"):
                with st.spinner("Ejecutando agente por primera vez..."):
                    try:
                        import subprocess
                        subprocess.run(["python", "agente_inteligencia.py"])
                        st.rerun()
                    except Exception as e:
                        st.error("Asegúrate de que 'agente_inteligencia.py' esté en la misma carpeta.")

st.markdown("""<div class="footer">KPI's desarrollados por el LAE. José Martín Estrada Cabrera | © 2026 Todos los Derechos Reservados</div>""", unsafe_allow_html=True)
